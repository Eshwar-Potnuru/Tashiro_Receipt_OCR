#!/usr/bin/env python3
"""Phase 6 pre-6B full system validation harness.

Runs end-to-end API flow with concurrent workers/admins and writes:
- artifacts/full_system_test_results.json
- artifacts/excel_smoke_verify_results.json (optional)
"""

from __future__ import annotations

import argparse
import json
import random
import statistics
import threading
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


@dataclass
class UserCred:
    email: str
    password: str
    role: str
    display_name: str


@dataclass
class ApiCall:
    actor: str
    endpoint: str
    method: str
    status_code: int
    elapsed_ms: float
    ok: bool
    error: str = ""
    headers: Optional[Dict[str, str]] = None


def _now_iso() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _read_users(seed_path: Path, worker_count: int, admin_count: int) -> Tuple[List[UserCred], List[UserCred]]:
    payload = json.loads(seed_path.read_text(encoding="utf-8"))

    workers: List[UserCred] = []
    admins: List[UserCred] = []
    for row in payload:
        email = row.get("email")
        if not email:
            continue
        cred = UserCred(
            email=email,
            password=row.get("password", ""),
            role=str(row.get("role", "")).upper(),
            display_name=row.get("display_name") or email,
        )
        if cred.role == "WORKER":
            workers.append(cred)
        elif cred.role == "ADMIN":
            admins.append(cred)

    if len(workers) < worker_count:
        raise ValueError(f"Need at least {worker_count} worker credentials in {seed_path}")
    if len(admins) < admin_count:
        raise ValueError(f"Need at least {admin_count} admin credentials in {seed_path}")

    return workers[:worker_count], admins[:admin_count]


def _load_default_staff_location(repo_root: Path) -> Tuple[Optional[str], Optional[str]]:
    locations_path = repo_root / "config" / "locations.json"
    staff_path = repo_root / "config" / "staff_config.json"
    if not locations_path.exists() or not staff_path.exists():
        return None, None

    locations_data = json.loads(locations_path.read_text(encoding="utf-8"))
    staff_data = json.loads(staff_path.read_text(encoding="utf-8"))

    for location in locations_data.get("locations", []):
        staff_list = staff_data.get(location) or []
        if staff_list:
            return location, staff_list[0].get("id")

    for location, staff_list in staff_data.items():
        if staff_list:
            return location, staff_list[0].get("id")
    return None, None


def _collect_receipt_files(receipt_root: Path) -> List[Path]:
    exts = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".heic"}
    files = [p for p in receipt_root.rglob("*") if p.is_file() and p.suffix.lower() in exts]
    return sorted(files)


def _timed_request(
    session: requests.Session,
    method: str,
    url: str,
    *,
    actor: str,
    endpoint: str,
    timeout: int = 60,
    **kwargs: Any,
) -> Tuple[Optional[requests.Response], ApiCall]:
    started = time.perf_counter()
    try:
        response = session.request(method=method, url=url, timeout=timeout, **kwargs)
        elapsed = (time.perf_counter() - started) * 1000
        slim_headers = {
            "X-Duplicate-Warning": response.headers.get("X-Duplicate-Warning", ""),
            "X-Tax-Mismatch-Warning": response.headers.get("X-Tax-Mismatch-Warning", ""),
            "X-OCR-Queue-Wait-Ms": response.headers.get("X-OCR-Queue-Wait-Ms", ""),
        }
        call = ApiCall(
            actor=actor,
            endpoint=endpoint,
            method=method,
            status_code=response.status_code,
            elapsed_ms=elapsed,
            ok=response.ok,
            error="" if response.ok else response.text[:300],
            headers=slim_headers,
        )
        return response, call
    except Exception as exc:
        elapsed = (time.perf_counter() - started) * 1000
        call = ApiCall(
            actor=actor,
            endpoint=endpoint,
            method=method,
            status_code=0,
            elapsed_ms=elapsed,
            ok=False,
            error=str(exc),
            headers=None,
        )
        return None, call


def _login(base_url: str, cred: UserCred) -> str:
    response = requests.post(
        f"{base_url}/api/auth/login",
        json={"email": cred.email, "password": cred.password},
        timeout=20,
    )
    if not response.ok:
        raise RuntimeError(f"Login failed for {cred.email}: {response.status_code} {response.text[:240]}")
    token = response.json().get("access_token")
    if not token:
        raise RuntimeError(f"No access token for {cred.email}")
    return token


def _p95(values: List[float]) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    idx = max(0, int(0.95 * len(ordered)) - 1)
    return round(ordered[idx], 2)


def _summarize_calls(calls: List[ApiCall]) -> Dict[str, Dict[str, float]]:
    grouped: Dict[str, List[ApiCall]] = {}
    for item in calls:
        grouped.setdefault(item.endpoint, []).append(item)

    summary: Dict[str, Dict[str, float]] = {}
    for endpoint, items in grouped.items():
        latencies = [i.elapsed_ms for i in items]
        ok_count = sum(1 for i in items if i.ok)
        summary[endpoint] = {
            "count": len(items),
            "ok": ok_count,
            "errors": len(items) - ok_count,
            "error_rate": round((len(items) - ok_count) / max(1, len(items)), 4),
            "avg_ms": round(statistics.mean(latencies), 2),
            "p95_ms": _p95(latencies),
            "max_ms": round(max(latencies), 2),
        }
    return summary


def _admin_poll_loop(
    base_url: str,
    token: str,
    actor: str,
    stop_event: threading.Event,
    poll_interval: float,
    all_calls: List[ApiCall],
    lock: threading.Lock,
) -> None:
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    targets = [
        ("GET", "/api/drafts", "GET /api/drafts"),
        ("GET", "/api/drafts/sent-records/all", "GET /api/drafts/sent-records/all"),
        ("GET", "/api/drafts/ledger-preview?format=staff", "GET /api/drafts/ledger-preview?format=staff"),
    ]

    while not stop_event.is_set():
        for method, path, endpoint in targets:
            _, call = _timed_request(
                session,
                method,
                f"{base_url}{path}",
                actor=actor,
                endpoint=endpoint,
                timeout=45,
            )
            with lock:
                all_calls.append(call)
        time.sleep(poll_interval)


def _edit_draft_receipt(
    session: requests.Session,
    base_url: str,
    actor: str,
    draft_id: str,
    fallback_location: Optional[str],
    fallback_staff_id: Optional[str],
) -> Tuple[bool, List[ApiCall]]:
    calls: List[ApiCall] = []
    draft_resp, call_get = _timed_request(
        session,
        "GET",
        f"{base_url}/api/drafts/{draft_id}",
        actor=actor,
        endpoint="GET /api/drafts/{id}",
    )
    calls.append(call_get)
    if draft_resp is None or not draft_resp.ok:
        return False, calls

    payload = draft_resp.json()
    receipt = payload.get("receipt") or {}

    if not receipt.get("receipt_date"):
        receipt["receipt_date"] = datetime.now().date().isoformat()
    if not receipt.get("vendor_name"):
        receipt["vendor_name"] = "Phase6 Upload Vendor"
    if not receipt.get("invoice_number"):
        receipt["invoice_number"] = f"UP-{uuid.uuid4().hex[:8]}"
    if receipt.get("total_amount") is None:
        receipt["total_amount"] = "1800"
    if receipt.get("tax_10_amount") is None:
        receipt["tax_10_amount"] = "900"
    if receipt.get("tax_8_amount") is None:
        receipt["tax_8_amount"] = "900"
    if not receipt.get("business_location_id") and fallback_location:
        receipt["business_location_id"] = fallback_location
    if not receipt.get("staff_id") and fallback_staff_id:
        receipt["staff_id"] = fallback_staff_id

    receipt["memo"] = f"phase6-pre6b-edit-{actor}-{draft_id[:8]}"

    update_resp, call_put = _timed_request(
        session,
        "PUT",
        f"{base_url}/api/drafts/{draft_id}",
        actor=actor,
        endpoint="PUT /api/drafts/{id}",
        json={"receipt": receipt},
    )
    calls.append(call_put)
    return bool(update_resp and update_resp.ok), calls


def _worker_flow(
    *,
    base_url: str,
    token: str,
    actor: str,
    files: List[Path],
    ocr_engine: str,
    stagger_seconds: float,
    fallback_location: Optional[str],
    fallback_staff_id: Optional[str],
) -> Dict[str, Any]:
    if stagger_seconds > 0:
        time.sleep(stagger_seconds)

    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    calls: List[ApiCall] = []
    created_draft_ids: List[str] = []
    upload_queue_wait_ms: Optional[int] = None

    opened = []
    multipart = []
    for file_path in files:
        fh = file_path.open("rb")
        opened.append(fh)
        multipart.append(("files", (file_path.name, fh, "application/octet-stream")))

    try:
        upload_resp, upload_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/batch-upload",
            actor=actor,
            endpoint="POST /api/drafts/batch-upload",
            files=multipart,
            data={"ocr_engine": ocr_engine},
            timeout=240,
        )
        calls.append(upload_call)
    finally:
        for fh in opened:
            try:
                fh.close()
            except Exception:
                pass

    if upload_call.headers and upload_call.headers.get("X-OCR-Queue-Wait-Ms"):
        try:
            upload_queue_wait_ms = int(upload_call.headers["X-OCR-Queue-Wait-Ms"])
        except Exception:
            upload_queue_wait_ms = None

    if upload_resp is not None and upload_resp.ok:
        body = upload_resp.json()
        for row in body.get("results", []):
            if row.get("status") == "success" and row.get("draft_id"):
                created_draft_ids.append(str(row["draft_id"]))

    edited_ok = 0
    for draft_id in created_draft_ids:
        success, edit_calls = _edit_draft_receipt(
            session,
            base_url,
            actor,
            draft_id,
            fallback_location,
            fallback_staff_id,
        )
        calls.extend(edit_calls)
        if success:
            edited_ok += 1

    _, list_call = _timed_request(
        session,
        "GET",
        f"{base_url}/api/drafts",
        actor=actor,
        endpoint="GET /api/drafts",
    )
    calls.append(list_call)

    send_headers: Dict[str, str] = {}
    sent_count = 0
    send_failed = 0
    if created_draft_ids:
        send_resp, send_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor=actor,
            endpoint="POST /api/drafts/send",
            json={"draft_ids": created_draft_ids},
            timeout=180,
        )
        calls.append(send_call)
        if send_call.headers:
            send_headers = send_call.headers
        if send_resp is not None:
            try:
                send_payload = send_resp.json()
                sent_count = int(send_payload.get("sent", 0))
                send_failed = int(send_payload.get("failed", 0))
            except Exception:
                pass

    return {
        "actor": actor,
        "uploaded_files": [f.name for f in files],
        "created_draft_ids": created_draft_ids,
        "upload_queue_wait_ms": upload_queue_wait_ms,
        "edited_ok": edited_ok,
        "send_headers": send_headers,
        "sent_count": sent_count,
        "send_failed": send_failed,
        "calls": calls,
    }


def _build_minimal_receipt(reference_receipt: Optional[Dict[str, Any]], *, mutate: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    today = datetime.now().date().isoformat()
    base = {
        "receipt_date": today,
        "vendor_name": "Phase6 Harness Vendor",
        "invoice_number": f"INV-{uuid.uuid4().hex[:8]}",
        "total_amount": "1200",
        "tax_10_amount": "600",
        "tax_8_amount": "600",
        "memo": "phase6-pre6b",
        "business_location_id": None,
        "staff_id": None,
    }
    if reference_receipt:
        base["business_location_id"] = reference_receipt.get("business_location_id")
        base["staff_id"] = reference_receipt.get("staff_id")
    if mutate:
        base.update(mutate)
    return base


def _extract_error_code(detail: Any) -> Optional[str]:
    if isinstance(detail, dict):
        return detail.get("error_code")
    return None


def _run_policy_checks(base_url: str, worker_token: str, seed_reference_receipt: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {worker_token}"})
    checks: Dict[str, Any] = {}
    policy_calls: List[ApiCall] = []

    def save_draft(receipt_payload: Dict[str, Any], image_ref: Optional[str] = None) -> Optional[str]:
        save_resp, save_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts",
            actor="policy_worker",
            endpoint="POST /api/drafts",
            json={
                "receipt": receipt_payload,
                "image_ref": image_ref or f"policy-{uuid.uuid4().hex[:10]}",
            },
        )
        policy_calls.append(save_call)
        if save_resp is None or not save_resp.ok:
            return None
        return str((save_resp.json() or {}).get("draft_id"))

    # Duplicate warning-only check
    duplicate_receipt = _build_minimal_receipt(
        seed_reference_receipt,
        mutate={
            "vendor_name": "Phase6 Duplicate Vendor",
            "invoice_number": "DUP-INV-001",
            "receipt_date": datetime.now().date().isoformat(),
            "total_amount": "2100",
            "tax_10_amount": "1000",
            "tax_8_amount": "1100",
        },
    )
    first_dup_id = save_draft(duplicate_receipt)
    second_dup_id = save_draft(duplicate_receipt)
    if first_dup_id:
        _, send_first_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [first_dup_id]},
        )
        policy_calls.append(send_first_call)

    duplicate_header = ""
    if second_dup_id:
        second_resp, send_second_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [second_dup_id]},
        )
        policy_calls.append(send_second_call)
        if send_second_call.headers:
            duplicate_header = send_second_call.headers.get("X-Duplicate-Warning", "")
        checks["duplicate_warning_only"] = {
            "expected": "header X-Duplicate-Warning=true and send not blocked",
            "status_code": send_second_call.status_code,
            "header": duplicate_header,
            "pass": send_second_call.status_code == 200 and duplicate_header.lower() == "true",
            "response_preview": second_resp.text[:220] if second_resp is not None else "",
        }
    else:
        checks["duplicate_warning_only"] = {
            "expected": "header X-Duplicate-Warning=true and send not blocked",
            "status_code": 0,
            "header": "",
            "pass": False,
            "response_preview": "Could not create duplicate test draft",
        }

    # Tax mismatch warning-only check
    tax_receipt = _build_minimal_receipt(
        seed_reference_receipt,
        mutate={
            "vendor_name": "Phase6 Tax Vendor",
            "invoice_number": f"TAX-{uuid.uuid4().hex[:6]}",
            "total_amount": "1000",
            "tax_10_amount": "700",
            "tax_8_amount": "500",
        },
    )
    tax_id = save_draft(tax_receipt)
    if tax_id:
        tax_resp, tax_send_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [tax_id]},
        )
        policy_calls.append(tax_send_call)
        tax_header = ""
        if tax_send_call.headers:
            tax_header = tax_send_call.headers.get("X-Tax-Mismatch-Warning", "")
        checks["tax_warning_only"] = {
            "expected": "header X-Tax-Mismatch-Warning=true and send not blocked",
            "status_code": tax_send_call.status_code,
            "header": tax_header,
            "pass": tax_send_call.status_code == 200 and tax_header.lower() == "true",
            "response_preview": tax_resp.text[:220] if tax_resp is not None else "",
        }
    else:
        checks["tax_warning_only"] = {
            "expected": "header X-Tax-Mismatch-Warning=true and send not blocked",
            "status_code": 0,
            "header": "",
            "pass": False,
            "response_preview": "Could not create tax mismatch test draft",
        }

    # Staff/location mismatch hard block (6A-4)
    reference_location = (seed_reference_receipt or {}).get("business_location_id")
    reference_staff_id = (seed_reference_receipt or {}).get("staff_id")
    mismatch_location = "Aichi" if reference_location != "Aichi" else "Tokyo"
    mismatch_staff = reference_staff_id or "tok_001"

    mismatch_receipt = _build_minimal_receipt(
        seed_reference_receipt,
        mutate={
            "vendor_name": "Phase6 Mismatch Vendor",
            "invoice_number": f"MIS-{uuid.uuid4().hex[:6]}",
            "business_location_id": mismatch_location,
            "staff_id": mismatch_staff,
        },
    )
    mismatch_id = save_draft(mismatch_receipt)
    mismatch_code = None
    mismatch_detail = ""
    mismatch_status = 0
    if mismatch_id:
        mismatch_resp, mismatch_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [mismatch_id]},
        )
        policy_calls.append(mismatch_call)
        mismatch_status = mismatch_call.status_code
        if mismatch_resp is not None:
            try:
                body = mismatch_resp.json()
                detail = body.get("detail")
                mismatch_code = _extract_error_code(detail)
                mismatch_detail = str(detail)[:220]
            except Exception:
                mismatch_detail = mismatch_resp.text[:220]
    checks["staff_location_mismatch_block"] = {
        "expected": "409 with STAFF_LOCATION_MISMATCH",
        "status_code": mismatch_status,
        "error_code": mismatch_code,
        "detail_preview": mismatch_detail,
        "pass": mismatch_status == 409 and mismatch_code == "STAFF_LOCATION_MISMATCH",
    }

    # SENT immutability block (6A-1) via re-send
    immutable_receipt = _build_minimal_receipt(
        seed_reference_receipt,
        mutate={
            "vendor_name": "Phase6 Immutable Vendor",
            "invoice_number": f"IMM-{uuid.uuid4().hex[:6]}",
        },
    )
    immutable_id = save_draft(immutable_receipt)
    immutable_status = 0
    immutable_detail = ""
    if immutable_id:
        _, first_send_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [immutable_id]},
        )
        policy_calls.append(first_send_call)

        second_resp, second_send_call = _timed_request(
            session,
            "POST",
            f"{base_url}/api/drafts/send",
            actor="policy_worker",
            endpoint="POST /api/drafts/send",
            json={"draft_ids": [immutable_id]},
        )
        policy_calls.append(second_send_call)
        immutable_status = second_send_call.status_code
        if second_resp is not None:
            immutable_detail = second_resp.text[:220]

    checks["sent_immutability_block"] = {
        "expected": "409 on re-send of already SENT draft",
        "status_code": immutable_status,
        "detail_preview": immutable_detail,
        "pass": immutable_status == 409,
    }

    checks["_calls"] = [c.__dict__ for c in policy_calls]
    return checks


def _run_excel_smoke_verification(repo_root: Path, since_epoch: float, output_path: Path) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "generated_at": _now_iso(),
        "since_epoch": since_epoch,
        "checked_files": [],
        "summary": {
            "total": 0,
            "valid_zip": 0,
            "openpyxl_loadable": 0,
            "failures": 0,
        },
    }

    accumulation_root = repo_root / "app" / "Data" / "accumulation"
    files = sorted(accumulation_root.rglob("*.xlsx")) if accumulation_root.exists() else []

    fresh_files = [p for p in files if p.stat().st_mtime >= since_epoch]
    targets = fresh_files if fresh_files else files

    for path in targets:
        entry: Dict[str, Any] = {
            "path": str(path.relative_to(repo_root)).replace("\\", "/"),
            "valid_zip": False,
            "openpyxl_loadable": False,
            "sheet_count": 0,
            "sheet_names": [],
            "error": "",
            "mtime_epoch": path.stat().st_mtime,
        }
        result["summary"]["total"] += 1
        try:
            entry["valid_zip"] = zipfile.is_zipfile(path)
            if entry["valid_zip"]:
                result["summary"]["valid_zip"] += 1
            if load_workbook is not None:
                wb = load_workbook(path, read_only=True, data_only=True)
                entry["sheet_names"] = list(wb.sheetnames)
                entry["sheet_count"] = len(wb.sheetnames)
                wb.close()
                entry["openpyxl_loadable"] = True
                result["summary"]["openpyxl_loadable"] += 1
            else:
                entry["error"] = "openpyxl unavailable; skipped workbook load"
        except Exception as exc:
            entry["error"] = str(exc)

        if not entry["valid_zip"] or (load_workbook is not None and not entry["openpyxl_loadable"]):
            result["summary"]["failures"] += 1
        result["checked_files"].append(entry)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Phase 6 pre-6B full flow validation harness")
    parser.add_argument("--base-url", default="http://127.0.0.1:8001")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--admins", type=int, default=1)
    parser.add_argument("--receipts-per-worker", type=int, default=2)
    parser.add_argument("--stagger-seconds", type=float, default=1.0)
    parser.add_argument("--poll-interval", type=float, default=1.0)
    parser.add_argument("--tail-poll-seconds", type=int, default=15)
    parser.add_argument("--ocr-engine", default="auto")
    parser.add_argument("--seed-users", default="config/users_seed_dev.json")
    parser.add_argument("--receipt-root", default="Sample reciepts")
    parser.add_argument("--output", default="artifacts/full_system_test_results.json")
    parser.add_argument("--excel-smoke-output", default="artifacts/excel_smoke_verify_results.json")
    parser.add_argument("--skip-excel-smoke", action="store_true")
    parser.add_argument("--p95-threshold-ms", type=float, default=2000.0)
    parser.add_argument("--error-rate-threshold", type=float, default=0.05)
    parser.add_argument("--strict", action="store_true", help="Return non-zero exit code on gate failure")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    output_path = repo_root / args.output
    excel_output_path = repo_root / args.excel_smoke_output
    seed_users_path = (repo_root / args.seed_users).resolve()
    receipt_root = (repo_root / args.receipt_root).resolve()

    suite_started_epoch = time.time()
    suite_started_iso = _now_iso()

    workers, admins = _read_users(seed_users_path, args.workers, args.admins)
    fallback_location, fallback_staff_id = _load_default_staff_location(repo_root)
    receipt_files = _collect_receipt_files(receipt_root)
    if len(receipt_files) < args.receipts_per_worker:
        raise RuntimeError(
            f"Not enough sample receipts in {receipt_root} for receipts_per_worker={args.receipts_per_worker}"
        )

    print(f"[1/6] Logging in {len(workers)} workers + {len(admins)} admins...")
    worker_tokens: List[str] = []
    admin_tokens: List[str] = []
    for user in workers:
        worker_tokens.append(_login(args.base_url, user))
    for user in admins:
        admin_tokens.append(_login(args.base_url, user))

    print("[2/6] Starting admin polling threads...")
    all_calls: List[ApiCall] = []
    lock = threading.Lock()
    stop_event = threading.Event()
    admin_threads: List[threading.Thread] = []
    for index, token in enumerate(admin_tokens, start=1):
        actor = f"admin_{index}"
        thread = threading.Thread(
            target=_admin_poll_loop,
            args=(args.base_url, token, actor, stop_event, args.poll_interval, all_calls, lock),
            daemon=True,
        )
        thread.start()
        admin_threads.append(thread)

    print("[3/6] Running worker upload/edit/send flows in parallel...")
    worker_outputs: List[Dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=len(workers)) as pool:
        futures = []
        for i, token in enumerate(worker_tokens, start=1):
            actor = f"worker_{i}"
            picked = random.sample(receipt_files, args.receipts_per_worker)
            futures.append(
                pool.submit(
                    _worker_flow,
                    base_url=args.base_url,
                    token=token,
                    actor=actor,
                    files=picked,
                    ocr_engine=args.ocr_engine,
                    stagger_seconds=(i - 1) * args.stagger_seconds,
                    fallback_location=fallback_location,
                    fallback_staff_id=fallback_staff_id,
                )
            )

        for future in as_completed(futures):
            output = future.result()
            worker_outputs.append(output)
            with lock:
                all_calls.extend(output["calls"])
            print(
                f"  {output['actor']}: drafts={len(output['created_draft_ids'])} "
                f"edited={output['edited_ok']} sent={output['sent_count']} failed={output['send_failed']}"
            )

    if args.tail_poll_seconds > 0:
        print(f"[4/6] Continuing admin polling for {args.tail_poll_seconds}s...")
        time.sleep(args.tail_poll_seconds)

    stop_event.set()
    for thread in admin_threads:
        thread.join(timeout=3)

    # Pick one worker and reference receipt for policy checks
    reference_receipt = None
    policy_token = worker_tokens[0]
    probe_session = requests.Session()
    probe_session.headers.update({"Authorization": f"Bearer {policy_token}"})
    probe_resp, probe_call = _timed_request(
        probe_session,
        "GET",
        f"{args.base_url}/api/drafts",
        actor="policy_worker",
        endpoint="GET /api/drafts",
    )
    all_calls.append(probe_call)
    if probe_resp is not None and probe_resp.ok:
        drafts = probe_resp.json() or []
        for row in drafts:
            rec = row.get("receipt") or {}
            if rec.get("staff_id") and rec.get("business_location_id"):
                reference_receipt = rec
                break

    print("[5/6] Running 6A policy checks (immutability, mismatch, duplicate, tax)...")
    policy_checks = _run_policy_checks(args.base_url, policy_token, reference_receipt)
    for row in policy_checks.get("_calls", []):
        all_calls.append(ApiCall(**row))

    endpoint_summary = _summarize_calls(all_calls)
    send_calls = [c for c in all_calls if c.endpoint == "POST /api/drafts/send"]
    normal_send_calls = [
        c
        for c in send_calls
        if c.actor.startswith("worker_")
    ]
    normal_send_ok = sum(1 for c in normal_send_calls if c.status_code == 200)

    non_policy_calls = [c for c in all_calls if not c.actor.startswith("policy_")]
    non_policy_unexpected_errors = [
        c for c in non_policy_calls if c.status_code == 0 or c.status_code >= 500
    ]
    gate_a = (len(non_policy_unexpected_errors) / max(1, len(non_policy_calls))) <= args.error_rate_threshold
    gate_b = True
    for critical in ["GET /api/drafts", "GET /api/drafts/sent-records/all"]:
        if critical in endpoint_summary and endpoint_summary[critical]["p95_ms"] > args.p95_threshold_ms:
            gate_b = False

    total_created = sum(len(item["created_draft_ids"]) for item in worker_outputs)
    total_sent = sum(int(item["sent_count"]) for item in worker_outputs)
    send_success_ratio = round(total_sent / max(1, total_created), 4)
    gate_c = send_success_ratio >= 0.95
    gate_d = all(
        bool(v.get("pass"))
        for k, v in policy_checks.items()
        if not k.startswith("_") and isinstance(v, dict)
    )

    excel_smoke_result = None
    gate_e = True
    if not args.skip_excel_smoke:
        excel_smoke_result = _run_excel_smoke_verification(repo_root, suite_started_epoch, excel_output_path)
        gate_e = (excel_smoke_result.get("summary", {}).get("failures", 0) == 0)

    suite_pass = gate_a and gate_b and gate_c and gate_d and gate_e

    payload = {
        "meta": {
            "suite": "phase6_pre6b_full_system_validation",
            "started_at": suite_started_iso,
            "finished_at": _now_iso(),
            "base_url": args.base_url,
        },
        "config": {
            "workers": args.workers,
            "admins": args.admins,
            "receipts_per_worker": args.receipts_per_worker,
            "stagger_seconds": args.stagger_seconds,
            "poll_interval": args.poll_interval,
            "tail_poll_seconds": args.tail_poll_seconds,
            "ocr_engine": args.ocr_engine,
            "receipt_root": str(receipt_root),
            "seed_users": str(seed_users_path),
            "thresholds": {
                "p95_ms": args.p95_threshold_ms,
                "error_rate": args.error_rate_threshold,
            },
        },
        "workers": [
            {
                "actor": item["actor"],
                "uploaded_files": item["uploaded_files"],
                "created_count": len(item["created_draft_ids"]),
                "edited_ok": item["edited_ok"],
                "sent_count": item["sent_count"],
                "send_failed": item["send_failed"],
                "upload_queue_wait_ms": item["upload_queue_wait_ms"],
                "warnings": {
                    "duplicate": item["send_headers"].get("X-Duplicate-Warning"),
                    "tax_mismatch": item["send_headers"].get("X-Tax-Mismatch-Warning"),
                },
            }
            for item in sorted(worker_outputs, key=lambda x: x["actor"])
        ],
        "metrics": endpoint_summary,
        "policy_checks": {
            k: v for k, v in policy_checks.items() if not k.startswith("_")
        },
        "gates": {
            "A_api_stability": gate_a,
            "B_polling_latency": gate_b,
            "C_send_completeness": gate_c,
            "D_policy_behavior": gate_d,
            "E_excel_smoke_optional": gate_e,
            "overall_pass": suite_pass,
            "send_success_ratio": send_success_ratio,
            "normal_send_total": len(normal_send_calls),
            "normal_send_ok": normal_send_ok,
            "worker_created_total": total_created,
            "worker_sent_total": total_sent,
        },
        "excel_smoke_output": str(excel_output_path.relative_to(repo_root)).replace("\\", "/") if not args.skip_excel_smoke else None,
        "calls": [c.__dict__ for c in all_calls],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print("[6/6] Completed")
    print(f"- Results: {output_path}")
    if not args.skip_excel_smoke:
        print(f"- Excel smoke: {excel_output_path}")
    print(
        "- Gates: "
        f"A={gate_a} B={gate_b} C={gate_c} D={gate_d} E={gate_e} overall={suite_pass}"
    )

    if args.strict and not suite_pass:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
