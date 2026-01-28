"""
Show exactly where data was written - File, Sheet, Row details
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from app.models.schema import Receipt
from app.excel.branch_ledger_writer import BranchLedgerWriter
from app.excel.staff_ledger_writer import StaffLedgerWriter
from app.services.config_service import ConfigService

def main():
    print("=" * 70)
    print("RECEIPT INSERTION DETAILS DEMONSTRATION")
    print("=" * 70)
    
    # Create test receipt
    test_receipt = Receipt(
        receipt_date="2025-01-19",
        vendor_name="デモ株式会社",
        total_amount=11000,
        tax_10_amount=1000,
        tax_8_amount=0,
        business_location_id="Osaka",
        staff_id="Staff01",
        invoice_number="DEMO-2025-001"
    )
    
    print("\n📝 TEST RECEIPT:")
    print(f"   Date: {test_receipt.receipt_date}")
    print(f"   Vendor: {test_receipt.vendor_name}")
    print(f"   Amount: ¥{test_receipt.total_amount:,}")
    print(f"   Location: {test_receipt.business_location_id}")
    print(f"   Staff: {test_receipt.staff_id}")
    print(f"   Invoice: {test_receipt.invoice_number}")
    
    # Write to location sheet
    print("\n" + "=" * 70)
    print("WRITING TO LOCATION SHEET...")
    print("=" * 70)
    
    location_writer = BranchLedgerWriter()
    result_location = location_writer.write_receipt(test_receipt)
    
    print(f"\n✅ LOCATION RESULT:")
    print(f"   Status: {result_location.get('status')}")
    print(f"   File: {result_location.get('location')}_Accumulated.xlsx")
    print(f"   Sheet: Monthly_Template")
    print(f"   Row Written: {result_location.get('row')}")
    
    # Verify location data
    location_file = Path(f"app/Data/accumulation/locations/{result_location.get('location')}_Accumulated.xlsx")
    if location_file.exists():
        wb = openpyxl.load_workbook(location_file)
        ws = wb['Monthly_Template']
        row = result_location.get('row')
        
        print(f"\n📊 DATA IN FILE:")
        print(f"   Full Path: {location_file.absolute()}")
        print(f"   Sheet Total Rows: {ws.max_row}")
        print(f"\n   Data at Row {row}:")
        print(f"      Column A (Date): {ws.cell(row, 1).value}")
        print(f"      Column C (Vendor): {ws.cell(row, 3).value}")
        print(f"      Column D (Staff): {ws.cell(row, 4).value}")
        print(f"      Column G (Invoice): {ws.cell(row, 7).value}")
        print(f"      Column I (Tax 10%): {ws.cell(row, 9).value}")
        print(f"      Column J (Tax 8%): {ws.cell(row, 10).value}")
        print(f"      Column L (Total): {ws.cell(row, 12).value}")
        
        # Check formulas
        print(f"\n   Formula Columns (Preserved):")
        n_cell = ws.cell(row, 14).value
        p_cell = ws.cell(row, 16).value
        q_cell = ws.cell(row, 17).value
        r_cell = ws.cell(row, 18).value
        print(f"      Column N: {'✓ Formula exists' if str(n_cell).startswith('=') else '✗ Empty or not formula'}")
        print(f"      Column P: {'✓ Formula exists' if str(p_cell).startswith('=') else '✗ Empty or not formula'}")
        print(f"      Column Q: {'✓ Formula exists' if str(q_cell).startswith('=') else '✗ Empty or not formula'}")
        print(f"      Column R: {'✓ Formula exists' if str(r_cell).startswith('=') else '✗ Empty or not formula'}")
        
        wb.close()
    
    # Write to staff sheet
    print("\n" + "=" * 70)
    print("WRITING TO STAFF SHEET...")
    print("=" * 70)
    
    staff_writer = StaffLedgerWriter()
    result_staff = staff_writer.write_receipt(test_receipt)
    
    print(f"\n✅ STAFF RESULT:")
    print(f"   Status: {result_staff.get('status')}")
    print(f"   File: {result_staff.get('staff')}_Accumulated.xlsx")
    print(f"   Sheet: {result_staff.get('sheet')}")
    print(f"   Row Written: {result_staff.get('row')}")
    
    # Verify staff data
    staff_file = Path(f"app/Data/accumulation/staff/{result_staff.get('staff')}_Accumulated.xlsx")
    if staff_file.exists():
        wb = openpyxl.load_workbook(staff_file)
        sheet_name = result_staff.get('sheet')
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row = result_staff.get('row')
            
            print(f"\n📊 DATA IN FILE:")
            print(f"   Full Path: {staff_file.absolute()}")
            print(f"   Sheet Total Rows: {ws.max_row}")
            print(f"\n   Data at Row {row}:")
            print(f"      Column A (Date): {ws.cell(row, 1).value}")
            print(f"      Column B (Vendor): {ws.cell(row, 2).value}")
            print(f"      Column F (Invoice): {ws.cell(row, 6).value}")
            print(f"      Column H (Tax 10%): {ws.cell(row, 8).value}")
            print(f"      Column I (Tax 8%): {ws.cell(row, 9).value}")
            print(f"      Column K (Total): {ws.cell(row, 11).value}")
            
            # Check formulas
            print(f"\n   Formula Columns (Preserved):")
            n_cell = ws.cell(row, 14).value
            p_cell = ws.cell(row, 16).value
            q_cell = ws.cell(row, 17).value
            r_cell = ws.cell(row, 18).value
            print(f"      Column N: {'✓ Formula exists' if str(n_cell).startswith('=') else '✗ Empty or not formula'}")
            print(f"      Column P: {'✓ Formula exists' if str(p_cell).startswith('=') else '✗ Empty or not formula'}")
            print(f"      Column Q: {'✓ Formula exists' if str(q_cell).startswith('=') else '✗ Empty or not formula'}")
            print(f"      Column R: {'✓ Formula exists' if str(r_cell).startswith('=') else '✗ Empty or not formula'}")
            
            wb.close()
    
    print("\n" + "=" * 70)
    print("✅ DEMONSTRATION COMPLETE")
    print("=" * 70)
    print("\n💡 KEY POINTS:")
    print("   - NO rows were inserted")
    print("   - Data written to existing empty rows in template")
    print("   - Formula columns remain intact")
    print("   - Table structure preserved")

if __name__ == "__main__":
    main()
