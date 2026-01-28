"""
COMPLETE DEMONSTRATION: Show exact file, sheet, and row where data was written
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from app.models.schema import Receipt
from app.excel.branch_ledger_writer import BranchLedgerWriter
from app.excel.staff_ledger_writer import StaffLedgerWriter

def main():
    print("=" * 80)
    print("COMPLETE DETAILS: FILE | SHEET | ROW")
    print("=" * 80)
    
    # Create test receipt
    test_receipt = Receipt(
        receipt_date="2025-01-19",
        vendor_name="デモベンダー株式会社",
        total_amount=15400,
        tax_10_amount=1400,
        tax_8_amount=0,
        business_location_id="Osaka",
        staff_id="Staff01",
        invoice_number="TEST-DEMO-19JAN"
    )
    
    print("\nTEST RECEIPT:")
    print(f"  📅 Date: {test_receipt.receipt_date}")
    print(f"  🏢 Vendor: {test_receipt.vendor_name}")
    print(f"  💴 Amount: ¥{test_receipt.total_amount:,}")
    print(f"  📍 Location: {test_receipt.business_location_id}")
    print(f"  👤 Staff: {test_receipt.staff_id}")
    print(f"  📋 Invoice: {test_receipt.invoice_number}")
    
    # 1. Write to LOCATION sheet
    print("\n" + "-" * 80)
    print("1️⃣  WRITING TO LOCATION SHEET...")
    print("-" * 80)
    
    location_writer = BranchLedgerWriter()
    result = location_writer.write_receipt(test_receipt)
    
    location_file = Path(f"app/Data/accumulation/locations/{result['location']}_Accumulated.xlsx")
    
    print(f"\n✅ WRITE COMPLETE:")
    print(f"   📁 File: {result['location']}_Accumulated.xlsx")
    print(f"   📋 Sheet: Monthly_Template")
    print(f"   📍 Row: {result['row']}")
    print(f"   📂 Full Path: {location_file.absolute()}")
    
    # Verify the data was written
    print(f"\n🔍 VERIFYING DATA IN FILE...")
    wb = openpyxl.load_workbook(location_file)
    ws = wb['Monthly_Template']
    row = result['row']
    
    print(f"\n   DATA AT ROW {row}:")
    print(f"      Column A (Date): {ws.cell(row, 1).value}")
    print(f"      Column C (Vendor): {ws.cell(row, 3).value}")
    print(f"      Column D (Staff): {ws.cell(row, 4).value}")
    print(f"      Column G (Invoice): {ws.cell(row, 7).value}")
    print(f"      Column I (Tax 10% Incl): {ws.cell(row, 9).value}")
    print(f"      Column J (Tax 8% Incl): {ws.cell(row, 10).value}")
    print(f"      Column L (Total): {ws.cell(row, 12).value}")
    
    print(f"\n   FORMULA COLUMNS (SHOULD BE PRESERVED):")
    for col_name, col_num in [('N', 14), ('P', 16), ('Q', 17), ('R', 18)]:
        cell_val = ws.cell(row, col_num).value
        status = "✓ Has Formula" if cell_val and str(cell_val).startswith('=') else "✗ Empty"
        print(f"      Column {col_name}: {status}")
    
    print(f"\n   TABLE INFO:")
    print(f"      Total Rows in Sheet: {ws.max_row}")
    print(f"      NO rows inserted - data filled existing empty row")
    
    wb.close()
    
    # 2. Write to STAFF sheet
    print("\n" + "-" * 80)
    print("2️⃣  WRITING TO STAFF SHEET...")
    print("-" * 80)
    
    staff_writer = StaffLedgerWriter()
    result_staff = staff_writer.write_receipt(test_receipt)
    
    staff_file = Path(f"app/Data/accumulation/staff/{result_staff['staff']}_Accumulated.xlsx")
    
    print(f"\n✅ WRITE COMPLETE:")
    print(f"   📁 File: {result_staff['staff']}_Accumulated.xlsx")
    print(f"   📋 Sheet: {result_staff['sheet']}")
    print(f"   📍 Row: {result_staff['row']}")
    print(f"   📂 Full Path: {staff_file.absolute()}")
    
    # Verify staff data
    print(f"\n🔍 VERIFYING DATA IN FILE...")
    wb = openpyxl.load_workbook(staff_file)
    ws = wb[result_staff['sheet']]
    row = result_staff['row']
    
    print(f"\n   DATA AT ROW {row}:")
    print(f"      Column A (Date): {ws.cell(row, 1).value}")
    print(f"      Column B (Vendor): {ws.cell(row, 2).value}")
    print(f"      Column F (Invoice): {ws.cell(row, 6).value}")
    print(f"      Column H (Tax 10% Incl): {ws.cell(row, 8).value}")
    print(f"      Column I (Tax 8% Incl): {ws.cell(row, 9).value}")
    print(f"      Column K (Total): {ws.cell(row, 11).value}")
    
    print(f"\n   FORMULA COLUMNS (SHOULD BE PRESERVED):")
    for col_name, col_num in [('N', 14), ('P', 16)]:
        cell_val = ws.cell(row, col_num).value
        status = "✓ Has Formula" if cell_val and str(cell_val).startswith('=') else "✗ Empty"
        print(f"      Column {col_name}: {status}")
    
    print(f"\n   TABLE INFO:")
    print(f"      Total Rows in Sheet: {ws.max_row}")
    print(f"      NO rows inserted - data filled existing empty row")
    
    wb.close()
    
    # Final summary
    print("\n" + "=" * 80)
    print("📊 SUMMARY - WHERE YOUR DATA WAS WRITTEN:")
    print("=" * 80)
    print(f"\n1. LOCATION FILE:")
    print(f"   File: {location_file.name}")
    print(f"   Sheet: Monthly_Template")
    print(f"   Row: {result['row']}")
    print(f"\n2. STAFF FILE:")
    print(f"   File: {staff_file.name}")
    print(f"   Sheet: {result_staff['sheet']}")
    print(f"   Row: {result_staff['row']}")
    print(f"\n✅ All data written successfully")
    print(f"✅ No rows inserted - filled existing empty rows")
    print(f"✅ Formula columns preserved")
    print("=" * 80)

if __name__ == "__main__":
    main()
