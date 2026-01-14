"""HQ Excel exporter for Jan 15 demo (Phase 2E demo scope).

Demo constraints (do not violate):
- Single receipt, single row insertion
- Expense-only
- Fixed templates; no structural changes; preserve merges
- No accumulation, no 集計シート logic, no carry-over
- No batch support

Responsibilities:
- Load template, copy to artifacts/demo_exports/
- Select target sheet by name
- Find first empty data row (format-specific start rows)
- Apply mappings defined in hq_excel_mappings
- Write values using fixed, source_field, or derived helpers

Not production-ready; intentionally minimal for demo.
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from app.exporters.hq_excel_mappings import (
    ColumnMapping,
    DERIVED_INVOICE_FLAG,
    DERIVED_TAX_10,
    DERIVED_TAX_8,
    DERIVED_TAX_EXEMPT,
    DERIVED_TAX_TOTAL_10,
    DERIVED_TAX_TOTAL_8,
    DERIVED_TAX_TOTAL_SUM,
    FORMAT_1_MAPPING,
    FORMAT_2_MAPPING,
)
from app.models.schema import ExtractionResult

# Paths
ROOT_DIR = Path(__file__).resolve().parents[2]
ARTIFACT_DIR = ROOT_DIR / "artifacts" / "demo_exports"
ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)


class HQExcelExporter:
    """Demo-only Excel exporter using predefined mappings.

    This exporter is intentionally narrow: single receipt, single row, no
    accumulation or pivot logic. It preserves the template structure and merged
    cells by copying the template before writing. Derived values are computed
    with simple helpers; anything unknown returns a safe default (None).
    """

    def __init__(self, template_path: Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

    def export_format_1(self, receipt: ExtractionResult, sheet_name: str) -> Path:
        """Export using Format ① (各個人集計用 _2024), writing one row to the given sheet."""
        return self._export_generic(
            receipt=receipt,
            sheet_name=sheet_name,
            mapping=FORMAT_1_MAPPING,
            first_data_row=2,
            header_row=1,
            format_label="format1"
        )

    def export_format_2(self, receipt: ExtractionResult, sheet_name: str) -> Path:
        """Export using Format ② (事業所集計テーブル), writing one row to the given sheet."""
        return self._export_generic(
            receipt=receipt,
            sheet_name=sheet_name,
            mapping=FORMAT_2_MAPPING,
            first_data_row=6,
            header_row=4,
            format_label="format2"
        )

    def _export_generic(
        self,
        receipt: ExtractionResult,
        sheet_name: str,
        mapping: Dict[str, ColumnMapping],
        first_data_row: int,
        header_row: int,
        format_label: str,
    ) -> Path:
        # Copy template to artifacts/demo_exports with timestamped name
        timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        out_path = ARTIFACT_DIR / f"{format_label}_{timestamp}.xlsx"
        shutil.copyfile(self.template_path, out_path)

        wb = load_workbook(out_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in template")
        ws = wb[sheet_name]

        # Determine first empty row starting at first_data_row
        target_row = None
        for r in range(first_data_row, ws.max_row + 2):
            if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1

        # Write mapped values
        header_to_col = _build_header_index(ws, header_row=header_row)
        for header, col_idx in header_to_col.items():
            if header not in mapping:
                continue
            mapping_def = mapping[header]
            value = self._resolve_value(receipt, mapping_def)
            # Write only if the mapping exists; None is allowed (will clear cell)
            ws.cell(row=target_row, column=col_idx, value=value)

        wb.save(out_path)
        return out_path

    def _resolve_value(self, receipt: ExtractionResult, mapping: ColumnMapping) -> Any:
        """Resolve a value based on fixed, source_field, or derived mapping."""
        if mapping.fixed_value is not None:
            return mapping.fixed_value
        if mapping.source_field:
            return getattr(receipt, mapping.source_field, None)
        if mapping.derived:
            return _derive_value(receipt, mapping.derived)
        return None


def _build_header_index(ws, header_row: int = 1) -> Dict[str, int]:
    """Build a header -> column index map using the provided header row."""
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val:
            headers[str(header_val)] = col
    return headers


def _derive_value(receipt: ExtractionResult, derived_key: str) -> Any:
    """Compute derived values for demo.

    - invoice_flag: "有" if invoice_number present, else "不要"
    - tax buckets route total based on tax_classification
    - tax totals are optional; default None when not computable
    """
    if derived_key == DERIVED_INVOICE_FLAG:
        return "有" if getattr(receipt, "invoice_number", None) else "不要"

    tax_class = getattr(receipt, "tax_classification", None)
    total = getattr(receipt, "total", None)

    if derived_key == DERIVED_TAX_10:
        return total if tax_class and "10" in str(tax_class) else None
    if derived_key == DERIVED_TAX_8:
        return total if tax_class and "8" in str(tax_class) else None
    if derived_key == DERIVED_TAX_EXEMPT:
        return total if tax_class and ("非課税" in str(tax_class) or "免税" in str(tax_class)) else None

    # Optional consumption tax outputs (demo-safe defaults)
    if derived_key == DERIVED_TAX_TOTAL_10:
        return None  # Could compute as total/1.1*0.1 if needed
    if derived_key == DERIVED_TAX_TOTAL_8:
        return None  # Could compute as total/1.08*0.08 if needed
    if derived_key == DERIVED_TAX_TOTAL_SUM:
        return None  # Could sum components if computed

    return None


__all__ = [
    "HQExcelExporter",
]
