"""Chronological insertion helpers (Phase 3A: no data writes)."""

from __future__ import annotations

from datetime import datetime
from typing import Optional, Sequence

from openpyxl.worksheet.worksheet import Worksheet


class InsertionRules:
    """Compute insertion positions and provide row-shift utilities.

    No receipt values are written in this phase.
    """

    def find_insertion_row(
        self,
        ws: Worksheet,
        receipt_date: Optional[str],
        *,
        date_column: int = 1,
        key_columns: Optional[Sequence[int]] = None,
        data_start_row: Optional[int] = None,
        footer_row: Optional[int] = None,
    ) -> int:
        """Return the row index to maintain chronological order.

        Defaults preserve existing Format 02 behavior. Callers may override:
        - date_column: column index containing the receipt date (default A/1).
        - key_columns: columns considered for detecting gaps before footer rows.
        - data_start_row: force the first data row when headers have no blank line.
        - footer_row: explicit footer/totals row to avoid appending below totals.
        """

        key_cols = tuple(key_columns) if key_columns else (1, 3, 4, 8, 11, 12, 14)

        if data_start_row is None:
            header_end = self._detect_header_end(ws)
            first_data_row = header_end + 1
        else:
            first_data_row = data_start_row
            header_end = first_data_row - 1

        detected_footer = self._detect_footer_start(ws, first_data_row, key_cols)
        footer_start = footer_row or detected_footer

        parsed_target = self._parse_date(receipt_date)
        max_row = footer_start - 1 if footer_start else ws.max_row

        if parsed_target is None:
            return min(max(max_row + 1, first_data_row), footer_start or (max_row + 1))

        insert_at = max(first_data_row, max_row + 1)
        for row in range(first_data_row, max_row + 1):
            cell_val = ws.cell(row=row, column=date_column).value
            parsed_cell = self._parse_date(cell_val)
            if parsed_cell is None:
                continue
            if parsed_target < parsed_cell:
                insert_at = row
                break
            if parsed_target == parsed_cell:
                insert_at = row + 1  # stable: after last same-date row
        return min(insert_at, footer_start or insert_at)

    def find_append_row_after_data(
        self,
        ws: Worksheet,
        *,
        key_columns: Sequence[int],
        first_data_row: int,
        footer_row: Optional[int] = None,
    ) -> int:
        """Return the row index immediately after the last data row (before footer).

        - key_columns: columns that indicate real transaction data (non-empty values).
        - first_data_row: where the data region starts.
        - footer_row: optional explicit footer start (e.g., 合計 row). Insertions stay above it.
        """

        max_scan_row = footer_row - 1 if footer_row else ws.max_row

        insert_row = first_data_row
        for row in range(max_scan_row, first_data_row - 1, -1):
            if self._row_has_data(ws, row, key_columns):
                insert_row = row + 1
                break

        if footer_row and insert_row > footer_row:
            insert_row = footer_row

        return insert_row

    def find_insert_row_before_footer(
        self,
        ws: Worksheet,
        *,
        data_start_row: int,
        key_columns: Sequence[int],
        footer_labels: Sequence[str],
    ) -> int:
        """Compute insertion row within a bounded data table before footer labels.

        Accumulator-style: find last row with real transaction data, insert after.
        Scans BACKWARD from footer to avoid template noise.

        - data_start_row: first row of the transaction table (e.g., 5)
        - key_columns: columns that indicate transaction data (non-empty)
        - footer_labels: strings that mark the footer (e.g., 合計, 残高)
        """

        # Find footer boundary - MUST stop here
        footer_row = self.find_footer_row_by_labels(
            ws,
            labels=footer_labels,
            start_row=data_start_row,
            columns=None,
        )
        
        if not footer_row:
            # No footer found - scan limited range
            max_scan = min(data_start_row + 40, ws.max_row)
        else:
            max_scan = footer_row - 1
        
        # Scan BACKWARD from max_scan to find last actual data row
        # This avoids counting template noise/empty formatted rows
        for row in range(max_scan, data_start_row - 1, -1):
            has_data = False
            for col in key_columns:
                val = ws.cell(row=row, column=col).value
                if val not in (None, "", " "):
                    has_data = True
                    break
            
            if has_data:
                # Found last data row, insert right after
                return row + 1
        
        # No data found at all - insert at data start
        return data_start_row

    def find_footer_row_by_labels(
        self,
        ws: Worksheet,
        *,
        labels: Sequence[str],
        start_row: int = 1,
        end_row: Optional[int] = None,
        columns: Optional[Sequence[int]] = None,
    ) -> Optional[int]:
        """Locate the first row containing any of the given labels.

        Useful for detecting totals/footer rows (e.g., a row labeled 合計) so
        callers can bound insertions above that row.
        """

        if end_row is None:
            end_row = ws.max_row
        cols = tuple(columns) if columns else tuple(range(1, ws.max_column + 1))

        target_tokens = {str(label).strip() for label in labels if label is not None}
        if not target_tokens:
            return None

        for row in range(start_row, end_row + 1):
            for col in cols:
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                text = str(val).strip()
                for token in target_tokens:
                    if token and token in text:
                        return row
        return None

    def shift_rows_preserve(self, ws: Worksheet, start_row: int, count: int = 1) -> None:
        """Shift rows down starting at start_row, preserving merges/styles/formulas.

        This uses openpyxl's built-in move_range; caller must handle value writes.
        """
        if count <= 0:
            return
        if start_row > ws.max_row:
            return
        max_col = ws.max_column or 1
        ws.move_range(
            f"A{start_row}:{self._col_letter(max_col)}{ws.max_row}",
            rows=count,
            cols=0,
            translate=True,
        )

    # -----------------
    # Helpers
    # -----------------
    @staticmethod
    def _parse_date(value: Optional[str]) -> Optional[datetime.date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.fromisoformat(str(value)).date()
        except Exception:
            return None

    @staticmethod
    def _detect_header_end(ws: Worksheet) -> int:
        """Detect the last header row by finding the first fully empty row from top."""
        row = 1
        while row <= ws.max_row:
            if InsertionRules._is_row_empty(ws, row):
                return row
            row += 1
        return ws.max_row

    @staticmethod
    def _detect_footer_start(ws: Worksheet, first_data_row: int, key_columns: Sequence[int]) -> Optional[int]:
        """Find the first gap that precedes totals/summary rows.

        We treat the first fully empty row within key columns (A, C, D, H, K, L, N)
        that has any content below it as the start of the footer region. This keeps
        insertions above totals and summary sections.
        """

        key_cols = tuple(key_columns)
        for row in range(first_data_row, ws.max_row + 1):
            if all(ws.cell(row=row, column=col).value in (None, "") for col in key_cols):
                if InsertionRules._has_content_below(ws, row + 1):
                    return row
        return None

    @staticmethod
    def _is_row_empty(ws: Worksheet, row: int) -> bool:
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return False
        return True

    @staticmethod
    def _has_content_below(ws: Worksheet, start_row: int) -> bool:
        max_row = ws.max_row
        max_col = ws.max_column or 1
        for row in range(start_row, max_row + 1):
            for col in range(1, max_col + 1):
                if ws.cell(row=row, column=col).value not in (None, ""):
                    return True
        return False

    @staticmethod
    def _row_has_data(ws: Worksheet, row: int, key_columns: Sequence[int]) -> bool:
        for col in key_columns:
            if ws.cell(row=row, column=col).value not in (None, ""):
                return True
        return False

    @staticmethod
    def _col_letter(idx: int) -> str:
        from openpyxl.utils import get_column_letter

        return get_column_letter(idx)
