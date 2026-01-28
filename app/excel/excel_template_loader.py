"""Excel template loader for Phase 3A (infrastructure only)."""

from __future__ import annotations

import shutil
import warnings
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelTemplateLoader:
    """Load base templates and materialize destination workbooks safely.

    No receipt data is written in this phase. Templates are never mutated in place.
    """

    def __init__(self) -> None:
        base_dir = Path(__file__).resolve().parents[2]
        self.templates_dir = base_dir / "Template" / "Formats"

        # Base templates
        self.format02_template = self.templates_dir / "事業所集計テーブル.xlsx"
        self.format01_template = self.templates_dir / "各個人集計用　_2024.xlsx"

        # Destinations
        data_root = base_dir / "app" / "Data" / "accumulation"
        self.location_dir = data_root / "locations"
        self.staff_dir = data_root / "staff"

        # Ensure base dirs exist (no files created yet)
        self.location_dir.mkdir(parents=True, exist_ok=True)
        self.staff_dir.mkdir(parents=True, exist_ok=True)

    # -----------------
    # Public API
    # -----------------
    def ensure_location_workbook(self, location_id: str) -> Path:
        """Return path to per-location workbook, copying from base if missing."""
        dest = self.location_dir / f"{location_id}_Accumulated.xlsx"
        if not dest.exists():
            self._copy_base(self.format02_template, dest)
        return dest

    def ensure_staff_workbook(self, staff_name: str, location_id: str, *, staff_id: Optional[str] = None) -> Path:
        """Return path to per-staff workbook, copying from base if missing.

        Canonical naming: {STAFF_NAME}_{LOCATION}.xlsx. If a legacy file
        (<staff_id>.xlsx) exists, copy it to the canonical name to preserve data.
        """

        safe_staff = self._safe_filename(staff_name or staff_id or "staff")
        safe_loc = self._safe_filename(location_id or "loc")
        dest = self.staff_dir / f"{safe_staff}_{safe_loc}.xlsx"

        legacy = None
        if staff_id:
            legacy = self.staff_dir / f"{staff_id}.xlsx"

        if not dest.exists():
            if legacy and legacy.exists():
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(legacy, dest)
            else:
                self._copy_base(self.format01_template, dest)
        return dest

    def ensure_format02_month_sheet(self, workbook, target_sheet: str) -> Worksheet:
        """Create a new month sheet ONLY from the canonical clean template.

        NEVER returns an existing sheet - always creates fresh from template.
        This prevents accidental writes to wrong/existing sheets.
        Caller must check for existing sheets before calling this.
        """

        if target_sheet in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_sheet}' already exists; will not duplicate.")

        template_wb, template_sheet = self._load_format02_template_sheet()
        new_ws = workbook.create_sheet(title=target_sheet)
        self._copy_sheet(template_sheet, new_ws)
        template_wb.close()
        return new_ws
    
    def create_month_sheet_from_template(self, workbook, target_sheet: str) -> Worksheet:
        """Create a clean month sheet from external template file.
        
        This is the ONLY way to create new month sheets. Never duplicates from
        existing sheets in the workbook to prevent data corruption.
        """
        if target_sheet in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_sheet}' already exists")
        
        # Load external template
        template_wb, template_sheet = self._load_format02_template_sheet()
        
        # Create new sheet in target workbook
        new_ws = workbook.create_sheet(title=target_sheet)
        
        # Copy structure from template
        self._copy_sheet(template_sheet, new_ws)
        
        # Close template workbook
        template_wb.close()
        
        return new_ws

    # -----------------
    # Helpers
    # -----------------
    def _copy_base(self, base_path: Path, dest_path: Path) -> None:
        if not base_path.exists():
            raise FileNotFoundError(f"Base template missing: {base_path}")
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(base_path, dest_path)

    def _load_format02_template_sheet(self):
        # Suppress openpyxl warnings about invalid pivot cache dependencies
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=".*invalid dependency definitions.*")
            wb = load_workbook(self.format02_template)
        
        # Prefer the dedicated "Monthly_Template" sheet
        if "Monthly_Template" in wb.sheetnames:
            return wb, wb["Monthly_Template"]
        
        # Fallback to any sheet with 年 and 月 in name
        for name in wb.sheetnames:
            if "年" in name and "月" in name:
                return wb, wb[name]
        
        # Last resort: active sheet
        return wb, wb.active

    def _copy_sheet(self, source: Worksheet, target: Worksheet) -> None:
        # Copy cell values and styles safely
        for row in source.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    # Skip placeholders inside merged ranges; merge copy handles them.
                    continue
                dst = target.cell(row=cell.row, column=cell.column)
                dst.value = cell.value
                
                # Copy individual style attributes instead of _style object to avoid index errors
                try:
                    if cell.font:
                        dst.font = cell.font.copy()
                    if cell.border:
                        dst.border = cell.border.copy()
                    if cell.fill:
                        dst.fill = cell.fill.copy()
                    if cell.alignment:
                        dst.alignment = cell.alignment.copy()
                    if cell.number_format:
                        dst.number_format = cell.number_format
                    if cell.protection:
                        dst.protection = cell.protection.copy()
                except Exception:
                    # If style copying fails, continue with just the value
                    pass

        # Copy merges
        for merged in source.merged_cells.ranges:
            target.merge_cells(str(merged))

        # Copy column widths
        for key, dim in source.column_dimensions.items():
            target.column_dimensions[key].width = dim.width

        # Copy row heights
        for idx, dim in source.row_dimensions.items():
            if dim.height:
                target.row_dimensions[idx].height = dim.height

    @staticmethod
    def _safe_filename(value: str) -> str:
        safe = "_".join(str(value).split())
        forbidden = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for ch in forbidden:
            safe = safe.replace(ch, "_")
        return safe

    def load_workbook(self, path: Path):
        """Load a workbook from disk (caller may choose read-only if desired)."""
        return load_workbook(path)

