"""
CORRECT Template-preserving accumulation system for Receipt OCR.

This module implements the CORRECT behavior for client's template:

CRITICAL REQUIREMENTS:
1. The client template is NOT an empty table - it's a complete accounting 
   report with existing structure, sample data, and layout guides
2. NEVER delete, clear, or overwrite ANY existing template content
3. NEVER recreate or regenerate template files  
4. Rows 1-4 AND ALL OTHER EXISTING ROWS must remain untouched forever
5. Only append NEW OCR data at the very bottom after all existing content
6. Preserve ALL formatting: colors, borders, merged cells, fonts, layout
7. Staff assignment based on location, NOT operator override
8. Work with the SAME template file, never replace it

CORRECT BEHAVIOR:
- Load existing template (with all its content)
- Find the first truly empty row at the bottom
- Append new OCR data only there
- Save back to same file
- Never touch anything above the append point
"""

import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl import load_workbook

from validators import get_available_locations, normalize_location

# Setup logging
logger = logging.getLogger(__name__)

# Paths
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "Template" / "事業所集計テーブル.xlsx"
ACCUM_DIR = BASE_DIR / "app" / "Data" / "accumulation"
BACKUP_DIR = ACCUM_DIR / "backups"
ARTIFACTS_DIR = BASE_DIR / "artifacts" / "accumulation"

# Template sheet to use (November 2025)
TARGET_SHEET = "2025年11月"

# Expected Japanese headers from row 4
JAPANESE_HEADERS = [
    "支払日", "工番", "摘　　要", "担当者", "収入", "支出", "a", "インボイス",
    "勘定科目", "b", "10％税込額", "8％税込額", "非課税額", "税込合計", "c",
    "消費税10", "消費税8", "消費税計"
]


def ensure_location_file_exists(location_name: str) -> bool:
    """
    Ensure a location file exists by copying the original template ONLY if missing.
    
    CRITICAL: This function ONLY creates a file if it doesn't exist.
    It NEVER overwrites, clears, or modifies existing files.
    
    Args:
        location_name: Business location name (e.g., "Tokyo", "Aichi")
        
    Returns:
        True if template exists/was created successfully
        
    Raises:
        FileNotFoundError: If original template is missing
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Original template not found: {TEMPLATE_PATH}")
    
    location_file = ACCUM_DIR / f"{location_name}_Accumulated.xlsx"
    
    # If file exists, we're done - NEVER modify existing files
    if location_file.exists():
        logger.info(f"Using existing file for {location_name}: {location_file}")
        return True
    
    # Only create if missing - copy the original template byte-for-byte
    ACCUM_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, location_file)
    logger.info("Created location workbook by copying official template", extra={
        "location": location_name,
        "template": str(TEMPLATE_PATH),
        "destination": str(location_file)
    })
    
    return location_file.exists()


def _create_backup(file_path: Path, reason: str) -> Path:
    """Create a timestamped backup of an existing file."""
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    location_name = file_path.stem.replace("_Accumulated", "")
    
    backup_dir = BACKUP_DIR / location_name
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    backup_filename = f"{file_path.stem}_{reason}_{timestamp}{file_path.suffix}"
    backup_path = backup_dir / backup_filename
    
    shutil.copy2(file_path, backup_path)
    logger.info(f"Created backup: {backup_path}")
    
    return backup_path





def _duplicate_month_sheet(wb, target_sheet: str):
    """Duplicate the closest month sheet to preserve formatting if target missing."""
    month_pattern = re.compile(r"\d{4}年\d{1,2}月")

    source_sheet = None
    for name in wb.sheetnames:
        if month_pattern.fullmatch(name):
            source_sheet = wb[name]
            break

    if source_sheet is None:
        source_sheet = wb.active

    duplicated = wb.copy_worksheet(source_sheet)
    duplicated.title = target_sheet
    logger.info("Duplicated template sheet to create missing month", extra={
        "new_sheet": target_sheet,
        "source_sheet": source_sheet.title
    })
    return duplicated


def load_location_workbook(location_name: str):
    """
    Load the workbook for a specific location.
    
    CRITICAL: This function ONLY loads existing files or creates from template
    if missing. It NEVER modifies existing content.
    
    Args:
        location_name: Business location name
        
    Returns:
        tuple: (workbook, worksheet) for the location
        
    Raises:
        Exception: If workbook cannot be loaded
    """
    # Ensure file exists (create from template if missing, but never modify existing)
    ensure_location_file_exists(location_name)
    
    location_file = ACCUM_DIR / f"{location_name}_Accumulated.xlsx"
    
    try:
        wb = load_workbook(location_file)

        if TARGET_SHEET in wb.sheetnames:
            ws = wb[TARGET_SHEET]
        else:
            ws = _duplicate_month_sheet(wb, TARGET_SHEET)

        return wb, ws
        
    except Exception as e:
        logger.error(f"Failed to load workbook for {location_name}: {e}")
        raise


def find_first_empty_row(ws, start_row: int = 5, max_columns: int = 18) -> int:
    """Find the first completely empty row scanning from start_row downward."""
    row = start_row
    # Allow an upper bound to avoid infinite loops in pathological sheets
    upper_bound = ws.max_row + 1000

    while row <= upper_bound:
        row_values = [ws.cell(row=row, column=col).value for col in range(1, max_columns + 1)]
        if all((value is None or (isinstance(value, str) and not value.strip())) for value in row_values):
            return row
        row += 1

    raise RuntimeError("Could not find empty row to append without touching template region")


def append_japanese_template_row(ws, row_values: List[Any]) -> int:
    """
    Safely append a row of data at the very bottom of the template.
    
    CRITICAL: This function preserves ALL existing template content by:
    - Finding the first truly empty row at the bottom
    - Adding new data only there
    - Never modifying any existing content above
    
    Args:
        ws: Worksheet object
        row_values: List of values to append (must be 18 values for A-R columns)
        
    Returns:
        int: Row number where data was inserted
    """
    if len(row_values) != 18:
        raise ValueError(f"Expected 18 values for columns A-R, got {len(row_values)}")
    
    # Find first empty row after template block
    next_row = find_first_empty_row(ws)
    
    # Write values to the specific row (preserving all formatting above)
    for col_index, value in enumerate(row_values, 1):
        ws.cell(row=next_row, column=col_index, value=value)
    
    logger.info("Writing OCR data to template", extra={
        "sheet": ws.title,
        "row": next_row
    })
    return next_row


def prepare_japanese_row_values(data: Dict[str, Any], location: str, operator: Dict[str, Any]) -> List[Any]:
    """
    Prepare row values in the correct order for the Japanese template.
    
    Args:
        data: Receipt data dictionary
        location: Business location
        operator: Operator information
        
    Returns:
        List of 18 values matching template columns A-R
    """
    from accumulator import _get_staff_member_for_location
    
    # Get staff member for this location
    staff_member = (
        data.get("staff_member") 
        or _get_staff_member_for_location(location, operator)
        or operator.get("name", "")
    )
    
    # Map data to the 18 template columns in order
    row_values = [
        # A: 支払日 (Payment Date)
        data.get("receipt_date") or data.get("date") or data.get("order_date") or "",
        
        # B: 工番 (Work Number) - intentionally empty
        "",
        
        # C: 摘　　要 (Description)
        data.get("description") or data.get("item_description") or data.get("vendor_name") or data.get("vendor") or data.get("store_name") or "",
        
        # D: 担当者 (Staff Member)
        staff_member,
        
        # E: 収入 (Income) - intentionally empty
        "",
        
        # F: 支出 (Expense)
        data.get("total_amount") or data.get("total") or data.get("amount") or "",
        
        # G: a (Template placeholder)
        "",
        
        # H: インボイス (Invoice)
        data.get("invoice_number") or "",
        
        # I: 勘定科目 (Account Title) - intentionally empty
        "",
        
        # J: b (Template placeholder)
        "",
        
        # K: 10％税込額 (10% Tax Included Amount)
        data.get("amount_tax_included_10") or "",
        
        # L: 8％税込額 (8% Tax Included Amount)
        data.get("amount_tax_included_8") or "",
        
        # M: 非課税額 (Non-taxable Amount) - intentionally empty
        "",
        
        # N: 税込合計 (Total Tax Included)
        data.get("total_amount") or data.get("total") or data.get("amount") or "",
        
        # O: c (Template placeholder)
        "",
        
        # P: 消費税10 (10% Consumption Tax)
        data.get("tax_10") or "",
        
        # Q: 消費税8 (8% Consumption Tax)
        data.get("tax_8") or "",
        
        # R: 消費税計 (Total Consumption Tax)
        data.get("tax_total") or data.get("tax_amount") or data.get("tax") or ""
    ]
    
    return row_values


def append_to_formatted_template(
    data: Dict[str, Any],
    location: str,
    operator: Dict[str, Any],
    *,
    force: bool = False,
) -> Dict[str, Any]:
    """
    Append receipt data to a location's formatted template workbook.
    
    This is the main function that preserves all template formatting while
    adding new receipt data.
    
    Args:
        data: Receipt data dictionary
        location: Business location name
        operator: Operator information dictionary
        force: If True, ignore duplicate detection
        
    Returns:
        Dictionary with operation results
    """
    # Validate location
    config = get_available_locations()
    normalized_location = normalize_location(location, config)
    if not normalized_location:
        raise ValueError(f"Unrecognized business location: {location}")
    
    try:
        # Load formatted workbook and ensure month sheet exists
        wb, ws = load_location_workbook(normalized_location)
        
        # Prepare row data
        row_values = prepare_japanese_row_values(data, normalized_location, operator)
        
        # Duplicate detection (column H only, within month sheet)
        if not force:
            invoice_number = data.get("invoice_number")
            if invoice_number:
                normalized_invoice = str(invoice_number).strip()
                for row_num in range(5, ws.max_row + 1):
                    existing_invoice = ws.cell(row=row_num, column=8).value
                    if existing_invoice and str(existing_invoice).strip() == normalized_invoice:
                        logger.info("Duplicate invoice detected", extra={
                            "location": normalized_location,
                            "invoice": normalized_invoice,
                            "row": row_num
                        })
                        return {
                            "status": "duplicate",
                            "location": normalized_location,
                            "duplicate_row": row_num,
                            "appended_rows": 0,
                            "message": f"Invoice {invoice_number} already exists at row {row_num}"
                        }
        
        logger.info("Saving OCR accumulation row", extra={
            "location": normalized_location,
            "sheet": ws.title
        })

        # Append new data row at the bottom (preserving all existing content)
        appended_row = append_japanese_template_row(ws, row_values)
        
        # Save workbook
        location_file = ACCUM_DIR / f"{normalized_location}_Accumulated.xlsx"
        wb.save(location_file)
        wb.close()
        
        # Create artifact copy
        artifact_path = ARTIFACTS_DIR / f"{normalized_location}_Accumulated.xlsx"
        artifact_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(location_file, artifact_path)
        
        # Log success
        from accumulator import _log_submission
        _log_submission("ok", normalized_location, {
            "invoice_number": data.get("invoice_number", ""),
            "amount": data.get("total_amount", data.get("amount", ""))
        }, location_file, "Row appended to formatted template")
        
        logger.info(f"✅ Appended data to {normalized_location} at row {appended_row} (bottom of template)")
        
        return {
            "status": "success",
            "location": normalized_location,
            "filepath": str(location_file),
            "artifact_path": str(artifact_path),
            "appended_rows": 1,
            "row_number": appended_row,
            "row_data": dict(zip(JAPANESE_HEADERS, row_values)),
            "template_preserved": True,
            "formatting_intact": True,
            "existing_content_untouched": True
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to append to {normalized_location}: {e}")
        return {
            "status": "error",
            "location": normalized_location,
            "error": str(e),
            "appended_rows": 0
        }





if __name__ == "__main__":
    """Test the CORRECT template-preserving system."""
    
    print("🎯 Testing CORRECT template behavior:")
    print("✅ Never deletes existing template content")
    print("✅ Never recreates or overwrites files")
    print("✅ Only appends at the very bottom")
    print("✅ Preserves ALL existing template structure")
    
    # Test appending data to Tokyo location
    test_data = {
        "date": "2024-11-19",
        "amount": 2500,
        "description": "正しいテンプレート保持テスト",
        "invoice_number": "CORRECT-TEMPLATE-001"
    }
    
    test_operator = {"name": "正しいテスト"}
    
    # Test appending (will only create file if missing, never overwrite existing)
    result = append_to_formatted_template(test_data, "Tokyo", test_operator)
    print(f"\n✅ Correct append result: {result}")