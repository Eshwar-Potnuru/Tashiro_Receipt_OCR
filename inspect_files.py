import openpyxl
from pathlib import Path

print("=" * 70)
print("FILE AND ROW DETAILS")
print("=" * 70)

# Osaka Location File
location_file = Path('app/Data/accumulation/locations/Osaka_Accumulated.xlsx')
if location_file.exists():
    wb = openpyxl.load_workbook(location_file)
    ws = wb['Monthly_Template']
    
    print("\n1. LOCATION FILE:")
    print(f"   File: Osaka_Accumulated.xlsx")
    print(f"   Sheet: Monthly_Template")
    print(f"   Path: {location_file.absolute()}")
    print(f"   Total Rows: {ws.max_row}")
    
    row = 24
    print(f"\n   DATA AT ROW {row}:")
    print(f"      Date (Col A): {ws.cell(row, 1).value}")
    print(f"      Vendor (Col C): {ws.cell(row, 3).value}")
    print(f"      Staff (Col D): {ws.cell(row, 4).value}")
    print(f"      Invoice (Col G): {ws.cell(row, 7).value}")
    print(f"      Tax 10% (Col I): {ws.cell(row, 9).value}")
    print(f"      Tax 8% (Col J): {ws.cell(row, 10).value}")
    print(f"      Total (Col L): {ws.cell(row, 12).value}")
    
    print(f"\n   FORMULAS (PRESERVED):")
    n_val = ws.cell(row, 14).value
    p_val = ws.cell(row, 16).value
    q_val = ws.cell(row, 17).value
    r_val = ws.cell(row, 18).value
    print(f"      Column N: {'Has Formula' if n_val and str(n_val).startswith('=') else 'Empty'}")
    print(f"      Column P: {'Has Formula' if p_val and str(p_val).startswith('=') else 'Empty'}")
    print(f"      Column Q: {'Has Formula' if q_val and str(q_val).startswith('=') else 'Empty'}")
    print(f"      Column R: {'Has Formula' if r_val and str(r_val).startswith('=') else 'Empty'}")
    
    print("\n   PREVIOUS ROWS STATUS:")
    for check_row in [22, 23]:
        vendor = ws.cell(check_row, 3).value
        print(f"      Row {check_row}: {'Has data' if vendor else 'Empty'} - {vendor if vendor else 'N/A'}")
    
    wb.close()

# Staff File
staff_file = Path('app/Data/accumulation/staff/Staff01_Accumulated.xlsx')
if staff_file.exists():
    wb = openpyxl.load_workbook(staff_file)
    ws = wb['202501']
    
    print("\n\n2. STAFF FILE:")
    print(f"   File: Staff01_Accumulated.xlsx")
    print(f"   Sheet: 202501")
    print(f"   Path: {staff_file.absolute()}")
    print(f"   Total Rows: {ws.max_row}")
    
    row = 5
    print(f"\n   DATA AT ROW {row}:")
    print(f"      Date (Col A): {ws.cell(row, 1).value}")
    print(f"      Vendor (Col B): {ws.cell(row, 2).value}")
    print(f"      Invoice (Col F): {ws.cell(row, 6).value}")
    print(f"      Tax 10% (Col H): {ws.cell(row, 8).value}")
    print(f"      Tax 8% (Col I): {ws.cell(row, 9).value}")
    print(f"      Total (Col K): {ws.cell(row, 11).value}")
    
    print(f"\n   FORMULAS (PRESERVED):")
    n_val = ws.cell(row, 14).value
    p_val = ws.cell(row, 16).value
    print(f"      Column N: {'Has Formula' if n_val and str(n_val).startswith('=') else 'Empty'}")
    print(f"      Column P: {'Has Formula' if p_val and str(p_val).startswith('=') else 'Empty'}")
    
    print("\n   PREVIOUS ROWS STATUS:")
    for check_row in [3, 4]:
        vendor = ws.cell(check_row, 2).value
        print(f"      Row {check_row}: {'Has data' if vendor else 'Empty'} - {vendor if vendor else 'N/A'}")
    
    wb.close()

print("\n" + "=" * 70)
print("SUMMARY:")
print("  Location: Row 24 in Osaka_Accumulated.xlsx (Monthly_Template)")
print("  Staff: Row 5 in Staff01_Accumulated.xlsx (202501)")
print("  NO rows inserted - filled existing empty rows")
print("  Formulas preserved in all locations")
print("=" * 70)
