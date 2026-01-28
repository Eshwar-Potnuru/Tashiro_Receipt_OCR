#!/usr/bin/env python3
"""Debug row finding and appending functions."""

import openpyxl
import template_formatter

def debug_row_finding():
    """Debug the row finding and appending process."""
    
    print("🔍 DEBUGGING ROW FINDING AND APPENDING")
    print("=" * 50)
    
    # Test the row finding function directly
    tokyo_file = 'app/Data/accumulation/Tokyo_Accumulated.xlsx'
    wb = openpyxl.load_workbook(tokyo_file)
    ws = wb.active

    print(f'📊 Worksheet max_row: {ws.max_row}')
    print(f'📊 Worksheet max_column: {ws.max_column}')

    # Test the function
    next_empty = template_formatter.find_first_empty_row_at_bottom(ws)
    print(f'🎯 Next empty row found: {next_empty}')

    print('\n🔍 Checking content around that area:')
    for row_num in range(max(1, next_empty - 5), min(ws.max_row + 5, next_empty + 5)):
        has_content = False
        content_summary = []
        
        for col in range(1, 10):
            value = ws.cell(row=row_num, column=col).value
            if value and str(value).strip():
                has_content = True
                content_summary.append(f'{chr(64+col)}:{str(value)[:10]}')
        
        if has_content or row_num == next_empty:
            marker = ' 👈 NEXT' if row_num == next_empty else ''
            content_display = content_summary if has_content else ['[EMPTY]']
            print(f'Row {row_num:2d}: {content_display}{marker}')

    wb.close()

    print('\n🔍 Testing manual append...')
    # Try to append manually to see what happens
    result = template_formatter.append_to_formatted_template(
        {
            'date': '2024-11-19', 
            'amount': 9999, 
            'invoice_number': 'MANUAL-TEST-999', 
            'description': 'Manual Debug Test'
        },
        'Tokyo',
        {'name': 'Debug Tester'}
    )
    
    print(f'📊 Manual append result: {result.get("status")}')
    print(f'📊 Row number: {result.get("row_number")}')
    
    if result.get("status") == "success":
        print("✅ Append successful - checking where data ended up...")
        
        # Reload and check
        wb = openpyxl.load_workbook(tokyo_file)
        ws = wb.active
        
        target_row = result.get("row_number")
        if target_row:
            print(f'\n🔍 Checking row {target_row} for our data:')
            row_data = []
            for col in range(1, 10):
                value = ws.cell(row=target_row, column=col).value
                row_data.append(str(value) if value else '')
            
            print(f'Row {target_row} content: {row_data}')
            
            # Check invoice column specifically
            invoice_val = ws.cell(row=target_row, column=8).value
            amount_val = ws.cell(row=target_row, column=6).value
            
            print(f'Invoice (col H): {invoice_val}')
            print(f'Amount (col F): {amount_val}')
        
        wb.close()
    else:
        print(f'❌ Append failed: {result.get("error", "Unknown error")}')

if __name__ == "__main__":
    debug_row_finding()