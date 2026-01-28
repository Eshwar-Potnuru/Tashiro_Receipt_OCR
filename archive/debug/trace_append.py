#!/usr/bin/env python3
"""Trace the append process to see why there's a row number discrepancy."""

import openpyxl
import template_formatter
import logging

# Enable debug logging
logging.basicConfig(level=logging.DEBUG)

def trace_append_process():
    """Trace through the append process step by step."""
    
    print("🔍 TRACING APPEND PROCESS")
    print("=" * 50)
    
    # Load workbook before append
    tokyo_file = 'app/Data/accumulation/Tokyo_Accumulated.xlsx'
    
    print("📊 BEFORE APPEND:")
    wb = openpyxl.load_workbook(tokyo_file)
    ws = wb.active
    
    print(f"   Max row: {ws.max_row}")
    
    # Find next row manually
    next_row_manual = template_formatter.find_first_empty_row_at_bottom(ws)
    print(f"   Next row (manual check): {next_row_manual}")
    
    wb.close()
    
    # Now do the append and trace what happens
    print(f"\n🔍 DURING APPEND:")
    test_data = {
        'date': '2024-11-19', 
        'amount': 7777, 
        'invoice_number': 'TRACE-TEST-777', 
        'description': 'Trace Test'
    }
    
    result = template_formatter.append_to_formatted_template(
        test_data,
        'Tokyo',
        {'name': 'Trace Tester'}
    )
    
    print(f"\n📊 APPEND RESULT:")
    print(f"   Status: {result.get('status')}")
    print(f"   Row number: {result.get('row_number')}")
    
    # Check what's actually in that row
    print(f"\n📊 AFTER APPEND:")
    wb = openpyxl.load_workbook(tokyo_file)
    ws = wb.active
    
    print(f"   New max row: {ws.max_row}")
    
    target_row = result.get('row_number')
    if target_row:
        print(f"\n🔍 Checking row {target_row}:")
        row_content = []
        for col in range(1, 10):
            value = ws.cell(row=target_row, column=col).value
            row_content.append(str(value) if value else '')
        print(f"   Content: {row_content}")
        
        # Check specifically invoice and amount
        invoice = ws.cell(row=target_row, column=8).value
        amount = ws.cell(row=target_row, column=6).value
        print(f"   Invoice (H): {invoice}")
        print(f"   Amount (F): {amount}")
        
        # Check if these match our test data
        if str(invoice) == 'TRACE-TEST-777' and str(amount) == '7777':
            print("   ✅ Data matches what we sent")
        else:
            print("   ❌ Data doesn't match what we sent")
            print(f"   Expected invoice: TRACE-TEST-777, got: {invoice}")
            print(f"   Expected amount: 7777, got: {amount}")
    
    # Also check if there's data in the row we expected (21)
    print(f"\n🔍 Checking expected row {next_row_manual}:")
    expected_row_content = []
    for col in range(1, 10):
        value = ws.cell(row=next_row_manual, column=col).value
        expected_row_content.append(str(value) if value else '')
    print(f"   Content: {expected_row_content}")
    
    if any(cell for cell in expected_row_content if cell.strip()):
        print("   ❓ Expected row has content - function may have been called with stale worksheet")
    else:
        print("   ✅ Expected row is empty as expected")
    
    wb.close()

if __name__ == "__main__":
    trace_append_process()