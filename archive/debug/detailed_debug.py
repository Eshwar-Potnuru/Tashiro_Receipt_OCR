#!/usr/bin/env python3
"""Detailed debug of the row finding discrepancy."""

import openpyxl
from pathlib import Path

def detailed_row_debug():
    """Debug the exact row finding behavior step by step."""
    
    print("🔍 DETAILED ROW FINDING DEBUG")
    print("=" * 50)
    
    tokyo_file = 'app/Data/accumulation/Tokyo_Accumulated.xlsx'
    wb = openpyxl.load_workbook(tokyo_file)
    ws = wb.active
    
    print(f"📊 Excel reported max_row: {ws.max_row}")
    print(f"📊 Excel reported max_column: {ws.max_column}")
    
    # Manual scan to find actual last content row
    print("\n🔍 Manual scan for last content row:")
    actual_last_content = 0
    
    for row_num in range(1, 50):  # Check beyond max_row
        has_content = False
        content_cols = []
        
        for col_num in range(1, 20):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is not None and str(cell_value).strip():
                has_content = True
                content_cols.append(f"{chr(64+col_num)}:{str(cell_value)[:8]}")
        
        if has_content:
            actual_last_content = row_num
            print(f"Row {row_num:2d}: {content_cols[:3]}...")  # Show first 3
    
    print(f"\n📊 Actual last content row: {actual_last_content}")
    print(f"📊 Expected next row: {actual_last_content + 1}")
    
    # Now test our function
    print(f"\n🔍 Testing our row finding function...")
    
    # Import and test
    import template_formatter
    found_row = template_formatter.find_first_empty_row_at_bottom(ws)
    print(f"🎯 Our function returned: {found_row}")
    
    if found_row != actual_last_content + 1:
        print(f"❌ MISMATCH! Expected {actual_last_content + 1}, got {found_row}")
        
        # Debug the function step by step
        print("\n🔍 Debugging function logic:")
        last_content_row = 0
        
        for row_num in range(1, ws.max_row + 1):
            has_any_content = False
            
            for col_num in range(1, 20):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    has_any_content = True
                    break
            
            if has_any_content:
                last_content_row = row_num
        
        next_empty_row = max(last_content_row + 1, 5)
        print(f"Function logic: last_content={last_content_row}, next_empty={next_empty_row}")
        
        # The issue might be that ws.max_row is wrong or cached
        print(f"Function uses ws.max_row={ws.max_row}, but actual content goes to {actual_last_content}")
    else:
        print("✅ Function working correctly")
    
    wb.close()

if __name__ == "__main__":
    detailed_row_debug()