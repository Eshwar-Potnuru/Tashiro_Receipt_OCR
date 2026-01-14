"""Per-location Excel accumulation utilities for Receipt OCR.

TODO: Capture full data-contract assumptions and edge-case handling in docs/services before refactoring.
"""
from __future__ import annotations

import csv
import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, List

import pandas as pd
import openpyxl
from openpyxl import load_workbook

from validators import (
    get_available_locations,
    normalize_location,
    normalize_number,
    parse_date,
)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "app" / "Data"
ACCUM_DIR = DATA_DIR / "accumulation"
BACKUP_DIR = ACCUM_DIR / "backups"
LOG_DIR = DATA_DIR / "submission_logs"
LOG_FILE = LOG_DIR / "submission_log.csv"
LOG_HEADERS = [
    "timestamp",
    "location",
    "sheet",
    "row",
    "staff_member",
    "invoice_number",
    "file_path",
    "backup_path",
    "operator_name",
    "operator_email",
    "operator_id",
    "status",
    "message",
]
ARTIFACTS_ACCUM_DIR = BASE_DIR / "artifacts" / "accumulation"
TEMPLATE_PATH = BASE_DIR / "Template" / "事業所集計テーブル.xlsx"
STAFF_CONFIG_PATH = BASE_DIR / "staff_config.json"
logger = logging.getLogger(__name__)
MAX_BACKUPS_PER_LOCATION = 3
DATA_ENTRY_START_ROW = 41


def _record_log_entry(payload: Dict[str, Any]) -> None:
    """Append a structured entry to the submission log CSV."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    write_header = not LOG_FILE.exists()
    with LOG_FILE.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=LOG_HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerow(payload)


def _create_backup_snapshot(file_path: Path, location: str) -> Optional[Path]:
    """Create a timestamped backup and prune old backups for the location."""
    if not file_path.exists():
        return None

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    location_dir = BACKUP_DIR / location
    location_dir.mkdir(parents=True, exist_ok=True)
    backup_name = f"{file_path.stem}_{timestamp}{file_path.suffix}"
    backup_path = location_dir / backup_name

    try:
        shutil.copy2(file_path, backup_path)
    except Exception as exc:
        logger.warning("Failed to create backup for %s: %s", location, exc)
        return None

    _prune_old_backups(location_dir, file_path.stem, file_path.suffix)
    return backup_path


def _prune_old_backups(location_dir: Path, stem: str, suffix: str) -> None:
    """Keep only the most recent MAX_BACKUPS_PER_LOCATION backups per location."""
    try:
        backups = sorted(
            location_dir.glob(f"{stem}_*{suffix}"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
    except FileNotFoundError:
        return

    for obsolete in backups[MAX_BACKUPS_PER_LOCATION:]:
        try:
            obsolete.unlink()
        except OSError:
            logger.warning("Unable to remove old backup %s", obsolete)


def get_staff_for_location(location: str) -> List[Dict[str, Any]]:
    """Get staff list for a specific location from staff configuration."""
    try:
        with STAFF_CONFIG_PATH.open("r", encoding="utf-8") as f:
            staff_config = json.load(f)
        return staff_config.get(location, [])
    except (FileNotFoundError, json.JSONDecodeError):
        logger.warning(f"Staff config not found or invalid, returning empty list for {location}")
        return []


def validate_staff_member(location: str, staff_member: str) -> bool:
    """Validate that staff_member exists for the given location."""
    staff_list = get_staff_for_location(location)
    return any(staff["name"] == staff_member for staff in staff_list)


def _get_staff_member_for_location(location: str, staff_member: str) -> Optional[Dict[str, Any]]:
    """
    Get staff member details for a location.
    Used by template_formatter for compatibility.
    """
    try:
        staff_list = get_staff_for_location(location)
        for staff in staff_list:
            if staff['name'] == staff_member:
                return staff
        return None
    except Exception:
        return None


def ensure_location_workbook(location: str) -> Path:
    """Ensure per-location workbook exists by copying template if needed."""
    _ensure_directories()
    location_file_path = ACCUM_DIR / f"{location}_Accumulated.xlsx"
    
    if not location_file_path.exists():
        if not TEMPLATE_PATH.exists():
            raise FileNotFoundError(f"Template file not found: {TEMPLATE_PATH}")
        
        # Copy template to create new location workbook
        shutil.copyfile(TEMPLATE_PATH, location_file_path)
        logger.info(f"Created new location workbook by copying template: {location_file_path}")
    
    return location_file_path


def get_month_sheet(wb, year: int, month: int):
    """Retrieve or create month sheet by duplicating existing sheet."""
    sheet_name = f"{year}年{month:02d}月"
    
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    
    # Need to create the sheet by duplicating an existing month sheet
    # Find a template sheet to duplicate (prefer current month or any month sheet)
    template_sheet_name = None
    
    # Look for existing month sheets
    for name in wb.sheetnames:
        if "年" in name and "月" in name and name not in ["記入例", "English ver 記入例", "集計テーブル"]:
            template_sheet_name = name
            break
    
    if not template_sheet_name:
        # Fallback to active sheet if no month sheets found
        template_sheet_name = wb.active.title
    
    # Duplicate the template sheet
    source_sheet = wb[template_sheet_name]
    new_sheet = wb.copy_worksheet(source_sheet)
    new_sheet.title = sheet_name
    
    logger.info(f"Created new month sheet '{sheet_name}' by duplicating '{template_sheet_name}'")
    return new_sheet


def find_first_empty_row(ws, start_row: int = DATA_ENTRY_START_ROW) -> int:
    """Find first fully empty row (A-R) starting at the designated data region."""
    max_col = 18  # Column R
    search_ceiling = max(ws.max_row, start_row)

    for row_num in range(start_row, search_ceiling + 2):
        row_cells = [ws.cell(row=row_num, column=col).value for col in range(1, max_col + 1)]
        if all(cell in (None, "") for cell in row_cells):
            return row_num

    return search_ceiling + 1


def append_row(ws, row_idx: int, mapped_values: Dict[str, Any]) -> None:
    """Append mapped values to specific row using A-R column mapping."""
    # Column mapping A-R as per specification
    column_mapping = {
        1: mapped_values.get("支払日", ""),      # A: 支払日
        2: mapped_values.get("工番", ""),        # B: 工番  
        3: mapped_values.get("摘要", ""),        # C: 摘要
        4: mapped_values.get("担当者", ""),      # D: 担当者
        5: mapped_values.get("収入", ""),        # E: 収入
        6: mapped_values.get("支出", ""),        # F: 支出
        7: mapped_values.get("a", ""),          # G: a
        8: mapped_values.get("インボイス番号", ""), # H: インボイス番号
        9: mapped_values.get("勘定科目", ""),     # I: 勘定科目
        10: mapped_values.get("b", ""),         # J: b
        11: mapped_values.get("10％税込額", ""),  # K: 10％税込額
        12: mapped_values.get("8％税込額", ""),   # L: 8％税込額
        13: mapped_values.get("非課税額", ""),    # M: 非課税額
        14: mapped_values.get("税込合計", ""),    # N: 税込合計
        15: mapped_values.get("c", ""),         # O: c
        16: mapped_values.get("消費税10", ""),   # P: 消費税10
        17: mapped_values.get("消費税8", ""),    # Q: 消費税8
        18: mapped_values.get("消費税計", ""),   # R: 消費税計
    }
    
    for col, value in column_mapping.items():
        ws.cell(row=row_idx, column=col, value=value)


def check_duplicate(ws, invoice_value: str) -> bool:
    """Check for duplicate invoice number in column H of the worksheet."""
    if not invoice_value:
        return False
    
    # Check column H (8th column) for duplicates, starting from row 5
    for row in range(5, ws.max_row + 1):
        existing_value = ws.cell(row=row, column=8).value
        if existing_value and str(existing_value) == str(invoice_value):
            return True
    
    return False


def persist_wb(wb, file_path: Path) -> None:
    """Save workbook with retry logic for Windows file locking."""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(file_path)
            return
        except PermissionError:
            if attempt == max_retries - 1:
                # Final attempt: clean up any existing temp files first
                for temp_attempt in range(max_retries):
                    temp_path = file_path.with_suffix(f".tmp_{temp_attempt}.xlsx")
                    try:
                        temp_path.unlink(missing_ok=True)
                    except:
                        pass
                # Now create new temp
                temp_path = file_path.with_suffix(f".tmp_{attempt}.xlsx")
                wb.save(temp_path)
                
                try:
                    if file_path.exists():
                        file_path.unlink()
                    temp_path.replace(file_path)
                    return
                except Exception as e:
                    # Clean up temp file on failure
                    try:
                        temp_path.unlink(missing_ok=True)
                    except:
                        pass
                    logger.error(f"Failed to save file after {max_retries} attempts: {e}")
                    raise
            else:
                import time
                time.sleep(0.5)


def log_operation(
    location: str,
    sheet_name: str,
    row_number: int,
    staff_member: str,
    invoice_number: str,
    operator: Dict[str, Any],
    file_path: Path,
    *,
    status: str = "success",
    message: str = "",
    backup_path: Optional[Path] = None,
) -> None:
    """Log operation details and persist them to CSV for operator traceability."""
    timestamp = datetime.utcnow().isoformat()

    log_message = (
        f"Location: {location}\n"
        f"Sheet: {sheet_name}\n"
        f"Row: {row_number}\n"
        f"Staff Member: {staff_member}\n"
        f"Invoice: {invoice_number}\n"
        f"Exact accumulation file path: {file_path}\n"
        f"Operator: {operator.get('name', '')} ({operator.get('email', '')}, id={operator.get('id', '')})\n"
        f"Status: {status}\n"
        f"Message: {message}\n"
        f"Timestamp: {timestamp}"
    )

    logger.info(log_message)

    log_entry = {
        "timestamp": timestamp,
        "location": location,
        "sheet": sheet_name,
        "row": row_number,
        "staff_member": staff_member,
        "invoice_number": invoice_number,
        "file_path": str(file_path),
        "backup_path": str(backup_path) if backup_path else "",
        "operator_name": operator.get("name", ""),
        "operator_email": operator.get("email", ""),
        "operator_id": operator.get("id", operator.get("employee_id", "")),
        "status": status,
        "message": message,
    }
    _record_log_entry(log_entry)


def append_to_month_sheet(data: Dict[str, Any], location: str, staff_member: str, 
                         operator: Dict[str, Any], *, force: bool = False) -> Dict[str, Any]:
    """Main function to append receipt data to monthly sheet following exact client spec."""
    
    # Validate staff member
    if not validate_staff_member(location, staff_member):
        raise ValueError(f"Invalid staff member '{staff_member}' for location '{location}'")
    
    # Ensure location workbook exists
    file_path = ensure_location_workbook(location)
    
    # Load workbook
    wb = load_workbook(file_path)
    
    # Determine target month from receipt_date
    receipt_date = data.get("receipt_date", "")
    if receipt_date:
        try:
            date_obj = datetime.strptime(receipt_date, "%Y-%m-%d")
            year, month = date_obj.year, date_obj.month
        except ValueError:
            # Fallback to current date
            now = datetime.now()
            year, month = now.year, now.month
    else:
        now = datetime.now()
        year, month = now.year, now.month
    
    # Get or create month sheet
    ws = get_month_sheet(wb, year, month)
    sheet_name = ws.title
    
    # Check for duplicates
    invoice_number = data.get("invoice_number", "")
    if invoice_number and check_duplicate(ws, invoice_number) and not force:
        log_operation(
            location,
            sheet_name,
            0,
            staff_member,
            invoice_number,
            operator,
            file_path,
            status="duplicate",
            message="Duplicate invoice rejected",
        )
        logger.warning(f"Duplicate invoice rejected: {invoice_number}")
        wb.close()
        return {
            "success": False,
            "error": "Duplicate invoice number",
            "invoice_number": invoice_number,
            "file_path": str(file_path),
            "sheet": sheet_name
        }
    
    backup_path = _create_backup_snapshot(file_path, location)

    # Find first empty row
    row_number = find_first_empty_row(ws)
    
    # Prepare mapped values
    mapped_values = {
        "支払日": data.get("receipt_date", ""),
        "工番": "",
        "摘要": data.get("vendor_name", ""),
        "担当者": staff_member,
        "収入": "",
        "支出": data.get("total_amount", ""),
        "a": "",
        "インボイス番号": invoice_number,
        "勘定科目": "",
        "b": "",
        "10％税込額": data.get("tax_10", ""),
        "8％税込額": data.get("tax_8", ""),
        "非課税額": "",
        "税込合計": data.get("total_amount", ""),
        "c": "",
        "消費税10": data.get("tax_10", ""),
        "消費税8": data.get("tax_8", ""),
        "消費税計": data.get("tax_total", "")
    }
    
    # Append row
    append_row(ws, row_number, mapped_values)
    
    # Save workbook
    try:
        persist_wb(wb, file_path)
    finally:
        wb.close()
    
    # Log operation
    log_operation(
        location,
        sheet_name,
        row_number,
        staff_member,
        invoice_number,
        operator,
        file_path,
        backup_path=backup_path,
        status="success",
        message="Receipt appended",
    )
    
    if force and invoice_number:
        logger.info(f"Force append used for duplicate invoice: {invoice_number}")
    
    return {
        "success": True,
        "file_path": str(file_path),
        "sheet": sheet_name,
        "row": row_number,
        "staff_member": staff_member,
        "invoice_number": invoice_number,
        "backup_path": str(backup_path) if backup_path else None,
    }
def _ensure_directories() -> None:
    """Create necessary directories."""
    DATA_DIR.mkdir(exist_ok=True)
    ACCUM_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    LOG_DIR.mkdir(exist_ok=True)
    ARTIFACTS_ACCUM_DIR.mkdir(parents=True, exist_ok=True)


# Legacy functions for compatibility (simplified)
def test_template_system(location: str = "Aichi") -> Dict[str, Any]:
    """Test function to validate the new accumulation system."""
    try:
        staff_list = get_staff_for_location(location)
        file_path = ensure_location_workbook(location)
        
        wb = load_workbook(file_path)
        ws = wb.active
        next_row = find_first_empty_row(ws)
        
        return {
            "status": "success",
            "next_row": next_row,
            "staff_count": len(staff_list),
            "file_exists": file_path.exists(),
            "template_path": str(TEMPLATE_PATH)
        }
    except Exception as e:
        return {
            "status": "error", 
            "error": str(e)
        }


def append_to_location(data: Dict[str, Any], location: str, operator: Dict[str, Any], 
                      *, force: bool = False) -> Dict[str, Any]:
    """Legacy wrapper function for compatibility."""
    # Extract staff_member from data or use first available for location
    staff_member = data.get("staff_member")
    if not staff_member:
        staff_list = get_staff_for_location(location)
        if staff_list:
            staff_member = staff_list[0]["name"]
        else:
            staff_member = operator.get("name", "Unknown")
    
    try:
        result = append_to_month_sheet(data, location, staff_member, operator, force=force)
        return {
            "status": "success" if result["success"] else "duplicate",
            "location": location,
            "filepath": result["file_path"],
            "appended_rows": 1 if result["success"] else 0,
            **result
        }
    except Exception as e:
        logger.error(f"Failed to append to location {location}: {e}")
        return {
            "status": "error",
            "error": str(e),
            "location": location
        }
