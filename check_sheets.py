#!/usr/bin/env python3
"""Check what sheets exist in the Tokyo file."""

import openpyxl

def check_sheets():
    """Check what sheets exist in the Tokyo workbook."""
    
    print("🔍 CHECKING WORKSHEET SHEETS")
    print("=" * 50)
    
    tokyo_file = 'app/Data/accumulation/Tokyo_Accumulated.xlsx'
    wb = openpyxl.load_workbook(tokyo_file)
    
    print(f"📊 Available sheets: {wb.sheetnames}")
    print(f"📊 Active sheet name: {wb.active.title}")
    
    # Check what TARGET_SHEET our code is looking for
    TARGET_SHEET = "2025年11月"
    print(f"📊 Looking for target sheet: '{TARGET_SHEET}'")
    
    if TARGET_SHEET in wb.sheetnames:
        print(f"✅ Target sheet exists")
        ws = wb[TARGET_SHEET]
    else:
        print(f"❌ Target sheet not found")
        
        # Check for month sheets
        month_sheets = [s for s in wb.sheetnames if "2025年" in s and "月" in s]
        print(f"📊 Month sheets found: {month_sheets}")
        
        if month_sheets:
            ws = wb[month_sheets[0]]
            print(f"✅ Using first month sheet: {month_sheets[0]}")
        else:
            ws = wb.active
            print(f"✅ Using active sheet: {ws.title}")
    
    # Check content in the sheet we're actually using
    print(f"\n🔍 Content in sheet '{ws.title}':")
    print(f"   Max row: {ws.max_row}")
    print(f"   Max col: {ws.max_column}")
    
    # Show first few rows
    print(f"\n📄 First few rows:")
    for row_num in range(1, min(6, ws.max_row + 1)):
        row_content = []
        for col in range(1, min(6, ws.max_column + 1)):
            value = ws.cell(row=row_num, column=col).value
            if value:
                row_content.append(f"{chr(64+col)}:{str(value)[:10]}")
        
        if row_content:
            print(f"   Row {row_num}: {row_content}")
    
    # Check last few rows
    print(f"\n📄 Last few rows:")
    for row_num in range(max(1, ws.max_row - 4), ws.max_row + 1):
        row_content = []
        for col in range(1, min(10, ws.max_column + 1)):
            value = ws.cell(row=row_num, column=col).value
            if value:
                row_content.append(f"{chr(64+col)}:{str(value)[:10]}")
        
        if row_content:
            print(f"   Row {row_num}: {row_content}")
        else:
            print(f"   Row {row_num}: [EMPTY]")
    
    # Look for our test data in ALL sheets
    print(f"\n🔍 Looking for test data in all sheets...")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        found_data = False
        
        for row_num in range(1, sheet.max_row + 1):
            for col in range(1, min(10, sheet.max_column + 1)):
                value = sheet.cell(row=row_num, column=col).value
                if value and ("TRACE-TEST" in str(value) or "MANUAL-TEST" in str(value) or "CORRECT-TEMPLATE" in str(value)):
                    print(f"   Sheet '{sheet_name}' Row {row_num} Col {chr(64+col)}: {value}")
                    found_data = True
        
        if not found_data:
            print(f"   Sheet '{sheet_name}': No test data found")
    
    wb.close()

if __name__ == "__main__":
    check_sheets()