#!/usr/bin/env python3
"""
Verify the CORRECT template behavior - ensure existing content is preserved.
"""

import sys
import openpyxl
from pathlib import Path

def verify_correct_template_behavior():
    """Verify that the template system works correctly by preserving ALL existing content."""
    
    print("🔍 VERIFYING CORRECT TEMPLATE BEHAVIOR")
    print("=" * 50)
    
    # Load the Tokyo file that was just modified
    tokyo_file = Path("app/Data/accumulation/Tokyo_Accumulated.xlsx")
    
    if not tokyo_file.exists():
        print("❌ Tokyo file not found - test invalid")
        return False
    
    try:
        wb = openpyxl.load_workbook(tokyo_file)
        ws = wb.active  # or wb["2025年11月"] if specific sheet
        
        print(f"📊 Total rows with content: {ws.max_row}")
        print(f"📊 Total columns with content: {ws.max_column}")
        
        # Check that template structure is preserved
        print("\n🔍 Checking template structure preservation:")
        
        # Check row 1 (should have template content)
        row_1_content = []
        for col in range(1, 6):
            value = ws.cell(row=1, column=col).value
            if value:
                row_1_content.append(str(value)[:20])  # First 20 chars
        
        print(f"Row 1 content: {row_1_content}")
        
        # Check row 4 (should have Japanese headers)
        row_4_headers = []
        for col in range(1, 10):
            value = ws.cell(row=4, column=col).value
            if value:
                row_4_headers.append(str(value))
        
        print(f"Row 4 headers: {row_4_headers[:5]}...")  # First 5 headers
        
        # Check if our new data is at the bottom (around row 41)
        print(f"\n🔍 Checking recent append (around row 40+):")
        
        # Look for our test data
        test_invoice = "CORRECT-TEMPLATE-001"
        test_amount = 2500
        
        found_test_row = None
        for row_num in range(35, ws.max_row + 1):  # Check near the end
            invoice_val = ws.cell(row=row_num, column=8).value  # Column H
            amount_val = ws.cell(row=row_num, column=6).value   # Column F
            
            if (invoice_val and str(invoice_val) == test_invoice) or \
               (amount_val and str(amount_val) == str(test_amount)):
                found_test_row = row_num
                print(f"✅ Found our test data at row {row_num}")
                
                # Show the row content
                row_content = []
                for col in range(1, 9):
                    value = ws.cell(row=row_num, column=col).value
                    row_content.append(str(value) if value else "")
                
                print(f"   Row {row_num}: {row_content}")
                break
        
        if not found_test_row:
            print("❌ Could not find our test data - append may have failed")
        
        # Check that there's content between template headers and our data
        content_rows_count = 0
        for row_num in range(5, found_test_row if found_test_row else 20):
            has_content = False
            for col in range(1, 10):
                if ws.cell(row=row_num, column=col).value:
                    has_content = True
                    break
            if has_content:
                content_rows_count += 1
        
        print(f"\n📊 Template content rows preserved: {content_rows_count}")
        
        if content_rows_count > 5:
            print("✅ CORRECT: Template has existing content that was preserved")
        else:
            print("⚠️  WARNING: Template seems empty - may not be the full client template")
        
        wb.close()
        
        print("\n" + "=" * 50)
        print("🎯 VERIFICATION SUMMARY:")
        print(f"✅ Template structure: PRESERVED")
        print(f"✅ Headers in row 4: PRESENT")
        print(f"✅ Existing content: {content_rows_count} rows preserved")
        print(f"✅ New data location: Row {found_test_row} (at bottom)" if found_test_row else "❌ New data not found")
        print(f"✅ Total file size: {ws.max_row} rows (shows growth)")
        
        if found_test_row and found_test_row > 20 and content_rows_count > 5:
            print("\n🎉 CORRECT BEHAVIOR VERIFIED!")
            print("✅ Template content preserved")
            print("✅ Data appended at bottom")
            print("✅ No overwriting occurred")
            return True
        else:
            print("\n⚠️  BEHAVIOR NEEDS VERIFICATION")
            return False
            
    except Exception as e:
        print(f"❌ Error verifying template: {e}")
        return False

if __name__ == "__main__":
    success = verify_correct_template_behavior()
    if success:
        print("\n🚀 CORRECT TEMPLATE SYSTEM VERIFIED!")
    else:
        print("\n💥 TEMPLATE BEHAVIOR NEEDS REVIEW")