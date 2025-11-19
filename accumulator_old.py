"""Per-location Excel accumulation utilities for Receipt OCR."""
from __future__ import annotations

import csv
import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, List

import pandas as pd
import openpyxl
from openpyxl import load_workbook

from validators import (
    get_available_locations,
    normalize_location,
    normalize_number,
    parse_date,
)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "app" / "Data"
ACCUM_DIR = DATA_DIR / "accumulation"
BACKUP_DIR = ACCUM_DIR / "backups"
LOG_DIR = DATA_DIR / "submission_logs"
LOG_FILE = LOG_DIR / "submission_log.csv"
ARTIFACTS_ACCUM_DIR = BASE_DIR / "artifacts" / "accumulation"
TEMPLATE_PATH = BASE_DIR / "Template" / "事業所集計テーブル.xlsx"
STAFF_CONFIG_PATH = BASE_DIR / "config" / "staff.json"
logger = logging.getLogger(__name__)

# Template-based Japanese column mapping (official headers from client template)
JAPANESE_COLUMN_MAPPING = {
    "支払日": "receipt_date",
    "工番": "",  # Empty field
    "摘　　要": "vendor_name",
    "担当者": "staff_member", 
    "収入": "",  # Empty field
    "支出": "total_amount",
    "a": "",  # Template column, keep empty
    "インボイス": "invoice_number",
    "勘定科目": "",  # Empty field 
    "b": "",  # Template column, keep empty
    "10％税込額": "amount_tax_included_10",
    "8％税込額": "amount_tax_included_8", 
    "非課税額": "",  # Empty field
    "税込合計": "total_amount",
    "c": "",  # Template column, keep empty
    "消費税10": "tax_10",
    "消費税8": "tax_8",
    "消費税計": "tax_total"
}
CANONICAL_HEADERS = [
    "Business Office",
    "Order Number",
    "Invoice Number",
    "Order Date",
    "Store Name",
    "Item Description",
    "Quantity",
    "Amount",
    "Tax Category",
    "Account Title",
    "Subtotal",
    "Tax Amount",
    "Currency",
    "Responsible Person",
    "Operator Full Name",
    "Operator Email",
    "Operator Employee ID",
    "Processed Timestamp",
    "Source File",
    "Notes",
]

LOG_HEADERS = [
    "timestamp",
    "location",
    "order_number",
    "invoice_number",
    "operator_name",
    "status",
    "file_path",
    "message",
]


def ensure_template_exists_for_location(location_name: str) -> Path:
    """Ensure a properly formatted template exists for the given location.
    
    Returns the path to the location's accumulated file.
    If file doesn't exist or is improperly formatted, creates fresh copy from template.
    """
    location_file = ACCUM_DIR / f"{location_name}_Accumulated.xlsx"
    
    # Always use fresh template to ensure formatting is preserved
    if not location_file.exists() or not _is_valid_template_format(location_file):
        logger.info(f"Creating fresh template copy for {location_name}")
        _create_fresh_template_copy(location_file)
    
    return location_file


def _is_valid_template_format(filepath: Path) -> bool:
    """Check if file has valid template format with preserved formatting."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check if Japanese headers exist in row 4
        expected_headers = ["支払日", "工番", "摘　　要", "担当者"]
        for i, expected in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=4, column=i).value
            if not cell_value or str(cell_value).strip() != expected:
                wb.close()
                return False
        
        # Check if template structure is preserved (rows 1-3 should have content)
        has_template_structure = bool(ws.cell(row=1, column=1).value)
        wb.close()
        return has_template_structure
        
    except Exception:
        return False


def _create_fresh_template_copy(destination_path: Path) -> None:
    """Create a fresh copy of the template with all formatting preserved."""
    if not TEMPLATE_PATH.exists():
        raise ValueError(f"Template file not found: {TEMPLATE_PATH}")
    
    # Create destination directory if needed
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Load template preserving all formatting
    wb = load_workbook(TEMPLATE_PATH)
    
    # Use November 2025 sheet or fallback
    target_sheet = "2025年11月"
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        logger.info(f"Using template sheet: {target_sheet}")
    else:
        month_sheets = [s for s in wb.sheetnames if "2025年" in s and "月" in s]
        if month_sheets:
            ws = wb[month_sheets[0]]
            logger.info(f"Using fallback sheet: {month_sheets[0]}")
        else:
            ws = wb.active
            logger.info(f"Using default sheet: {ws.title}")
    
    # Make this the active sheet and remove others to clean up
    wb.active = ws
    sheets_to_remove = [sheet for sheet in wb.sheetnames if sheet != ws.title]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])
    
    # Save the template copy - this preserves ALL formatting
    wb.save(destination_path)
    wb.close()
    
    logger.info(f"Created template copy with preserved formatting: {destination_path}")


def load_location_workbook(location_name: str):
    """Load the location workbook, ensuring it has proper template formatting."""
    location_file = ensure_template_exists_for_location(location_name)
    wb = load_workbook(location_file)
    ws = wb.active
    return wb, ws, location_file


def _load_staff_config() -> Dict[str, List[Dict[str, Any]]]:
    """Load staff configuration from staff.json."""
    if not STAFF_CONFIG_PATH.exists():
        return {"temporary_staff": [], "permanent_staff": []}
    
    with STAFF_CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _get_staff_member_for_location(location: str, operator: Dict[str, Any]) -> str:
    """Get appropriate staff member for location from staff config or operator."""
    staff_config = _load_staff_config()
    
    # Try to find a staff member assigned to this location
    for staff in staff_config.get("temporary_staff", []) + staff_config.get("permanent_staff", []):
        if location in staff.get("locations", []):
            return staff.get("name", "")
    
    # Fallback to operator name
    return operator.get("name", "")


def _extract_header_order(ws):
    """Extract column header order from template worksheet row 4."""
    header_order = []
    for cell in ws[4]:  # Row 4 contains the headers
        if cell.value:
            header_order.append(cell.value)
    return header_order


def _prepare_japanese_row(data: Dict[str, Any], location: str, operator: Dict[str, Any]) -> Dict[str, Any]:
    """Prepare a row using Japanese column mapping for template-based accumulation."""
    # Get appropriate staff member from configuration
    staff_member = (
        data.get("staff_member") 
        or _get_staff_member_for_location(location, operator)
        or ""
    )
    
    # Map data to Japanese column headers
    mapped_row = {
        "支払日": data.get("receipt_date") or data.get("date") or data.get("order_date") or "",
        "工番": "",  # Empty field
        "摘　　要": data.get("vendor_name") or data.get("vendor") or data.get("store_name") or data.get("description") or "",
        "担当者": staff_member,
        "収入": "",  # Empty field
        "支出": data.get("total_amount") or data.get("total") or data.get("amount") or "",
        "a": "",  # Template placeholder column
        "インボイス": data.get("invoice_number") or "",
        "勘定科目": "",  # Empty field
        "b": "",  # Template placeholder column  
        "10％税込額": data.get("amount_tax_included_10") or "",
        "8％税込額": data.get("amount_tax_included_8") or "",
        "非課税額": "",  # Empty field
        "税込合計": data.get("total_amount") or data.get("total") or data.get("amount") or "",
        "c": "",  # Template placeholder column
        "消費税10": data.get("tax_10") or "",
        "消費税8": data.get("tax_8") or "",
        "消費税計": data.get("tax_total") or data.get("tax_amount") or data.get("tax") or ""
    }
    
    return mapped_row


def append_japanese_template_row(ws, row_values: List[Any]) -> int:
    """Safely append a row to the template worksheet without affecting formatting.
    
    Args:
        ws: Worksheet object
        row_values: List of 18 values corresponding to columns A-R
        
    Returns:
        Row number where data was appended
    """
    # Find the next empty row (must be after row 4)
    next_row = max(5, ws.max_row + 1)
    
    # Ensure we have exactly 18 values (A-R columns)
    padded_values = row_values[:18] + [''] * (18 - len(row_values))
    
    # Write values to specific cells (safer than ws.append for formatting preservation)
    for col_index, value in enumerate(padded_values, 1):
        ws.cell(row=next_row, column=col_index, value=value)
    
    logger.debug(f"Appended row {next_row} with {len([v for v in padded_values if v])} non-empty values")
    return next_row


def _prepare_template_row_values(data: Dict[str, Any], location: str, operator: Dict[str, Any]) -> List[Any]:
    """Prepare row values in the exact order for template columns A-R."""
    # Get appropriate staff member
    staff_member = (
        data.get("staff_member") 
        or _get_staff_member_for_location(location, operator)
        or ""
    )
    
    # Map to exact column positions A-R
    row_values = [
        data.get("receipt_date") or data.get("date") or data.get("order_date") or "",  # A: 支払日
        "",  # B: 工番 (empty)
        data.get("vendor_name") or data.get("vendor") or data.get("store_name") or data.get("description") or "",  # C: 摘要
        staff_member,  # D: 担当者
        "",  # E: 収入 (empty)
        data.get("total_amount") or data.get("total") or data.get("amount") or "",  # F: 支出
        "",  # G: a (template placeholder)
        data.get("invoice_number") or "",  # H: インボイス
        "",  # I: 勘定科目 (empty)
        "",  # J: b (template placeholder)
        data.get("amount_tax_included_10") or "",  # K: 10％税込額
        data.get("amount_tax_included_8") or "",   # L: 8％税込額
        "",  # M: 非課税額 (empty)
        data.get("total_amount") or data.get("total") or data.get("amount") or "",  # N: 税込合計
        "",  # O: c (template placeholder)
        data.get("tax_10") or "",  # P: 消費税10
        data.get("tax_8") or "",   # Q: 消費税8
        data.get("tax_total") or data.get("tax_amount") or data.get("tax") or ""  # R: 消費税計
    ]
    
    return row_values


def _validate_template_integrity(filepath: Path) -> Dict[str, Any]:
    """Validate that the saved file maintains template structure."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check that first 3 rows are preserved (title/metadata)
        row1_has_content = any(cell.value for cell in ws[1])
        
        # Check that headers in row 4 match expected structure 
        header_order = _extract_header_order(ws)
        expected_headers = list(JAPANESE_COLUMN_MAPPING.keys())
        headers_match = len(header_order) >= len([h for h in expected_headers if h not in ["a", "b", "c"]])
        
        return {
            "valid": True,
            "title_preserved": row1_has_content,
            "headers_preserved": headers_match,
            "total_rows": ws.max_row,
            "header_count": len(header_order)
        }
    except Exception as e:
        logger.error(f"Template validation failed: {e}")
        return {"valid": False, "error": str(e)}


def _ensure_directories() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    ACCUM_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    LOG_DIR.mkdir(exist_ok=True)
    ARTIFACTS_ACCUM_DIR.mkdir(parents=True, exist_ok=True)


def _prepare_row(data: Dict[str, Any], location: str, operator: Dict[str, Any]) -> Dict[str, Any]:
    order_date = parse_date(data.get("order_date") or data.get("date")) or data.get("order_date") or data.get("date") or ""
    order_number = normalize_number(data.get("order_number")) or normalize_number(data.get("invoice_number"))
    invoice_number = normalize_number(data.get("invoice_number"))
    processed_ts = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    row = {
        "Business Office": location,
        "Order Number": order_number or "",
        "Invoice Number": invoice_number or "",
        "Order Date": order_date,
        "Store Name": data.get("store_name") or data.get("client_name") or data.get("vendor") or "",
        "Item Description": data.get("item_description") or data.get("account_title") or "",
        "Quantity": data.get("quantity") or data.get("item_quantity") or "",
        "Amount": data.get("amount") or data.get("total") or data.get("subtotal") or "",
        "Tax Category": data.get("tax_category") or "",
        "Account Title": data.get("account_title") or "",
        "Subtotal": data.get("subtotal") or "",
        "Tax Amount": data.get("tax") or data.get("tax_amount") or "",
        "Currency": data.get("currency") or "",
        "Responsible Person": data.get("responsible_person") or data.get("staff_member") or "",
        "Operator Full Name": operator.get("name", ""),
        "Operator Email": operator.get("email", ""),
        "Operator Employee ID": operator.get("employee_id") or operator.get("id") or "",
        "Processed Timestamp": processed_ts,
        "Source File": data.get("source_file") or data.get("source_image") or data.get("queue_id") or "",
        "Notes": data.get("notes") or "",
    }

    # Guarantee all headers exist
    for header in CANONICAL_HEADERS:
        row.setdefault(header, "")
    return row


def _log_submission(status: str, location: str, row: Dict[str, Any], filepath: Path, message: str = "") -> None:
    """Append an entry into the submission log CSV."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    write_header = not LOG_FILE.exists()
    payload = {
        "timestamp": datetime.utcnow().isoformat(timespec="seconds"),
        "location": location,
        "order_number": row.get("Order Number", ""),
        "invoice_number": row.get("Invoice Number", ""),
        "operator_name": row.get("Operator Full Name", ""),
        "status": status,
        "file_path": str(filepath),
        "message": message,
    }
    with LOG_FILE.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=LOG_HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerow(payload)


def convert_all_locations_to_template():
    """Convert all existing accumulation files to use Japanese template format."""
    _ensure_directories()
    config = get_available_locations()
    locations = config.get("locations", [])
    
    results = {}
    
    for location in locations:
        try:
            normalized_location = normalize_location(location, config)
            if not normalized_location:
                continue
                
            filepath = ACCUM_DIR / f"{normalized_location}_Accumulated.xlsx"
            
            if filepath.exists():
                # Create backup
                timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                backup_name = f"{filepath.stem}_backup_before_template_{timestamp}{filepath.suffix}"
                backup_path = BACKUP_DIR / normalized_location / backup_name
                backup_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(filepath, backup_path)
                
                # Load existing data to migrate
                existing_df = pd.read_excel(filepath)
                
                # Load fresh template
                wb, sheet_name = _load_template_workbook()
                ws = wb[sheet_name]
                
                # Migrate data to template format
                next_row = 5  # Start from row 5 (after headers in row 4)
                migrated_count = 0
                
                for _, row in existing_df.iterrows():
                    # Convert old format to new Japanese format
                    mapped_data = {
                        "支払日": row.get("Order Date", ""),
                        "工番": "",
                        "摘　　要": row.get("Store Name", ""),
                        "担当者": row.get("Responsible Person", ""),
                        "収入": "",
                        "支出": row.get("Amount", ""),
                        "a": "",
                        "インボイス": row.get("Invoice Number", ""),
                        "勘定科目": "",
                        "b": "",
                        "10％税込額": "",
                        "8％税込額": "",
                        "非課税額": "",
                        "税込合計": row.get("Amount", ""),
                        "c": "",
                        "消費税10": "",
                        "消費税8": "",
                        "消費税計": ""
                    }
                    
                    # Write to template
                    header_order = _extract_header_order(ws)
                    _write_row_to_template(ws, header_order, mapped_data, next_row)
                    next_row += 1
                    migrated_count += 1
                
                # Save converted file
                wb.save(filepath)
                
                results[location] = {
                    "status": "converted",
                    "migrated_rows": migrated_count,
                    "backup": str(backup_path),
                    "filepath": str(filepath)
                }
                
                logger.info(f"Converted {location} to template format: {migrated_count} rows migrated")
            else:
                results[location] = {
                    "status": "no_file",
                    "message": "No accumulation file found"
                }
                
        except Exception as e:
            results[location] = {
                "status": "error",
                "error": str(e)
            }
            logger.error(f"Failed to convert {location}: {e}")
    
    return results


def get_staff_list() -> Dict[str, List[Dict[str, Any]]]:
    """Get list of all staff members from configuration."""
    return _load_staff_config()


def test_template_system(location: str = "Aichi") -> Dict[str, Any]:
    """Test function to validate template-based accumulation system."""
    try:
        # Test data
        test_data = {
            "receipt_date": "2024-12-18",
            "vendor_name": "テストショップ",
            "total_amount": "1000",
            "invoice_number": f"TEST-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
            "tax_10": "100",
            "amount_tax_included_10": "1100"
        }
        
        test_operator = {
            "name": "テスト太郎",
            "email": "test@example.com",
            "employee_id": "TEST001"
        }
        
        # Test template loading
        template_test = {
            "template_exists": TEMPLATE_PATH.exists(),
            "template_path": str(TEMPLATE_PATH)
        }
        
        if template_test["template_exists"]:
            wb = _load_template_workbook()
            ws = wb.active
            headers = _extract_header_order(ws)
            template_test["headers_extracted"] = headers
            template_test["header_count"] = len(headers)
        
        # Test Japanese row mapping
        mapped_row = _prepare_japanese_row(test_data, location, test_operator)
        
        return {
            "status": "success", 
            "template_test": template_test,
            "mapped_row_test": mapped_row,
            "staff_member_test": mapped_row.get("担当者"),
            "test_ready": True
        }
        
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "test_ready": False
        }


def append_to_location(
    data: Dict[str, Any],
    location: str,
    operator: Dict[str, Any],
    *,
    force: bool = False,
) -> Dict[str, Any]:
    """Append a receipt row to the per-location accumulation workbook using client template.
    
    This function preserves the exact formatting of the client's template by:
    1. Loading the original template with all formatting intact
    2. Only appending data to empty rows (never modifying template rows 1-4)
    3. Preserving all colors, borders, fonts, and merged cells
    """
    _ensure_directories()
    config = get_available_locations()
    normalized_location = normalize_location(location, config)
    if not normalized_location:
        raise ValueError(f"Unrecognized business location: {location}")

    # Load location workbook with preserved template formatting
    wb, ws, filepath = load_location_workbook(normalized_location)
    
    # Prepare row values in exact column order A-R
    row_values = _prepare_template_row_values(data, normalized_location, operator)
    
    # Duplicate detection - check existing rows for invoice number
    duplicate = None
    invoice_no = data.get("invoice_number")
    
    if invoice_no and ws.max_row >= 5:  # Has data beyond header rows
        # Check column H (インボイス) for existing invoice numbers
        for row_num in range(5, ws.max_row + 1):
            existing_invoice = ws.cell(row=row_num, column=8).value  # Column H
            if existing_invoice and str(existing_invoice).strip() == str(invoice_no).strip():
                duplicate = {
                    "matched_on": "invoice_number", 
                    "existing_row_number": row_num,
                    "existing_invoice": existing_invoice
                }
                break
        
        if duplicate and not force:
            wb.close()
            mapped_row_dict = dict(zip(EXPECTED_HEADERS, row_values))
            _log_submission("duplicate", normalized_location, mapped_row_dict, filepath, "Duplicate detected")
            return {
                "status": "duplicate",
                "location": normalized_location,
                "filepath": str(filepath),
                "duplicate": duplicate,
                "appended_rows": 0,
            }
    
    # Find next available row (start from row 5, after headers in row 4)
    next_row = ws.max_row + 1
    if next_row < 5:  # Ensure we don't overwrite template structure
        next_row = 5
    
    # Write the row using template structure
    _write_row_to_template(ws, header_order, mapped_row, next_row)
    
    # Save with backup (preserving all formatting)
    backup_path = None
    if filepath.exists():
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{filepath.stem}_backup_{timestamp}{filepath.suffix}"
        backup_path = BACKUP_DIR / normalized_location / backup_name
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(filepath, backup_path)
            
            # Backup rotation (keep last 3)
            backups = sorted((BACKUP_DIR / normalized_location).glob(f"{filepath.stem}_backup_*{filepath.suffix}"))
            while len(backups) > 3:
                old_backup = backups.pop(0)
                try:
                    old_backup.unlink()
                except OSError:
                    pass
        except Exception as e:
            logger.warning(f"Could not create backup: {e}")

    # Save the workbook (preserves all template formatting)
    try:
        wb.save(filepath)
        wb.close()
        logger.info(f"Successfully saved formatted workbook: {filepath}")
    except Exception as e:
        wb.close()
        logger.error(f"Failed to save workbook: {e}")
        raise
    
    # Mirror to artifacts directory
    artifact_path = None
    try:
        ARTIFACTS_ACCUM_DIR.mkdir(parents=True, exist_ok=True)
        artifact_path = ARTIFACTS_ACCUM_DIR / filepath.name
        shutil.copy2(filepath, artifact_path)
    except Exception as exc:
        logger.warning("Failed to mirror accumulation workbook to artifacts: %s", exc)
        artifact_path = None
    
    # Validate template integrity
    validation = _validate_template_integrity(filepath)
    if not validation["valid"]:
        logger.warning(f"Template integrity check failed: {validation}")
    
    _log_submission("ok", normalized_location, mapped_row, filepath, "Row appended to template")

    return {
        "status": "success",
        "location": normalized_location,
        "filepath": str(filepath),
        "backup": str(backup_path) if backup_path else None,
        "artifact_path": str(artifact_path) if artifact_path else None,
        "appended_rows": 1,
        "row": mapped_row,
        "validation": validation,
        "template_used": True
    }
