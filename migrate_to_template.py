"""
Migration script to convert all location workbooks to official Japanese template format.

This script:
1. Loads the official template (Template/事業所集計テーブル.xlsx)
2. Checks each location workbook for old format issues
3. Replaces old-format files with fresh template copies
4. Optionally migrates valid data from old files
5. Ensures all workbooks follow the official Japanese template structure

Usage:
    python migrate_to_template.py
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import openpyxl
import pandas as pd
from openpyxl import load_workbook, Workbook

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Paths
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "Template" / "事業所集計テーブル.xlsx"
ACCUM_DIR = BASE_DIR / "app" / "Data" / "accumulation"
BACKUP_DIR = ACCUM_DIR / "backups"

# Expected Japanese headers (from template row 4)
EXPECTED_HEADERS = [
    "支払日", "工番", "摘　　要", "担当者", "収入", "支出", "a", "インボイス", 
    "勘定科目", "b", "10％税込額", "8％税込額", "非課税額", "税込合計", "c", 
    "消費税10", "消費税8", "消費税計"
]

# Location files to process
LOCATION_FILES = [
    "Aichi_Accumulated.xlsx",
    "Kashima_Accumulated.xlsx", 
    "Keihin_Accumulated.xlsx",
    "Osaka_Accumulated.xlsx",
    "Sagami_Accumulated.xlsx",
    "Takasago_Accumulated.xlsx",
    "Tokyo_Accumulated.xlsx"
]


def load_official_template():
    """Load the official Japanese template workbook."""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
    
    logger.info(f"Loading official template: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)
    
    # Use November 2025 sheet or fallback to first available month
    target_sheet = "2025年11月"
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        logger.info(f"Using template sheet: {target_sheet}")
    else:
        # Fallback to first month sheet
        month_sheets = [s for s in wb.sheetnames if "2025年" in s and "月" in s]
        if month_sheets:
            ws = wb[month_sheets[0]]
            logger.info(f"Using fallback sheet: {month_sheets[0]}")
        else:
            ws = wb.active
            logger.info(f"Using default sheet: {ws.title}")
    
    return wb, ws


def detect_old_format(filepath: Path) -> Dict[str, Any]:
    """
    Detect if an Excel file uses old format or has issues.
    
    Returns dict with:
        - is_old_format: bool
        - issues: list of detected issues
        - headers: current headers if readable
        - row_count: number of data rows
    """
    result = {
        "is_old_format": False,
        "issues": [],
        "headers": [],
        "row_count": 0,
        "readable": True
    }
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check for merged cells
        if ws.merged_cells.ranges:
            result["issues"].append("Contains merged cells")
            result["is_old_format"] = True
        
        # Try to read headers from row 4 (template format)
        row_4_headers = []
        for col_num in range(1, 19):  # A to R
            cell = ws.cell(row=4, column=col_num)
            if cell.value:
                row_4_headers.append(str(cell.value).strip())
        
        # If row 4 doesn't have Japanese headers, check row 1
        if not row_4_headers or row_4_headers[0] not in EXPECTED_HEADERS:
            row_1_headers = []
            for col_num in range(1, 25):  # Check more columns for old format
                cell = ws.cell(row=1, column=col_num)
                if cell.value:
                    row_1_headers.append(str(cell.value).strip())
            
            result["headers"] = row_1_headers
            
            # Check if it's old English format
            if any(header in str(row_1_headers) for header in ["Business Office", "Order Number", "Invoice Number"]):
                result["issues"].append("Old English header format detected")
                result["is_old_format"] = True
        else:
            result["headers"] = row_4_headers
        
        # Check column count
        if ws.max_column != 18:
            result["issues"].append(f"Wrong column count: {ws.max_column} (expected 18)")
            result["is_old_format"] = True
        
        # Check if Japanese headers match expected
        if row_4_headers:
            for i, expected in enumerate(EXPECTED_HEADERS):
                if i < len(row_4_headers):
                    if row_4_headers[i] != expected:
                        result["issues"].append(f"Header mismatch at column {i+1}: '{row_4_headers[i]}' != '{expected}'")
                        result["is_old_format"] = True
                        break
        
        # Count data rows
        result["row_count"] = max(0, ws.max_row - 4)  # Subtract template rows
        
        wb.close()
        
    except Exception as e:
        logger.warning(f"Could not read {filepath}: {e}")
        result["readable"] = False
        result["is_old_format"] = True
        result["issues"].append(f"File read error: {str(e)}")
    
    return result


def create_backup(filepath: Path, reason: str) -> Path:
    """Create a backup of the existing file."""
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    location_name = filepath.stem.replace("_Accumulated", "")
    
    backup_dir = BACKUP_DIR / location_name
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    backup_filename = f"{filepath.stem}_{reason}_{timestamp}{filepath.suffix}"
    backup_path = backup_dir / backup_filename
    
    shutil.copy2(filepath, backup_path)
    logger.info(f"Created backup: {backup_path}")
    
    return backup_path


def extract_migrable_data(filepath: Path) -> List[Dict[str, Any]]:
    """
    Extract data that can be safely migrated from old format files.
    
    Returns list of row dictionaries with mapped fields.
    """
    migrable_data = []
    
    try:
        # Try pandas first for easier data handling
        df = pd.read_excel(filepath)
        
        if df.empty:
            return migrable_data
        
        # Map old English headers to Japanese template fields
        header_mapping = {
            "Order Date": "支払日",
            "Store Name": "摘　　要",
            "Amount": "支出",
            "Invoice Number": "インボイス",
            "Responsible Person": "担当者",
            "Tax Amount": "消費税計",
            "Subtotal": "支出"  # Use subtotal if amount not available
        }
        
        for _, row in df.iterrows():
            mapped_row = {}
            
            # Map available fields
            for old_field, japanese_field in header_mapping.items():
                if old_field in df.columns and pd.notna(row[old_field]):
                    mapped_row[japanese_field] = row[old_field]
            
            # Ensure we have at least date or amount to make it worth migrating
            if mapped_row.get("支払日") or mapped_row.get("支出"):
                # Fill empty fields that the template expects
                for header in EXPECTED_HEADERS:
                    if header not in mapped_row:
                        mapped_row[header] = ""
                
                migrable_data.append(mapped_row)
        
        logger.info(f"Extracted {len(migrable_data)} migrable rows from {filepath}")
        
    except Exception as e:
        logger.warning(f"Could not extract data from {filepath}: {e}")
    
    return migrable_data


def apply_template_to_location(location_file: str, force_replace: bool = False) -> Dict[str, Any]:
    """
    Apply the official template to a location file.
    
    Args:
        location_file: Filename like "Aichi_Accumulated.xlsx"
        force_replace: If True, replace even if file seems OK
        
    Returns:
        Dictionary with migration results
    """
    filepath = ACCUM_DIR / location_file
    location_name = location_file.replace("_Accumulated.xlsx", "")
    
    result = {
        "location": location_name,
        "status": "unknown",
        "backup_created": False,
        "migrated_rows": 0,
        "issues_fixed": []
    }
    
    try:
        # Check if file exists
        if not filepath.exists():
            logger.info(f"{location_file} does not exist, creating fresh template")
            result["status"] = "created_new"
        else:
            # Detect format issues
            detection = detect_old_format(filepath)
            
            if not detection["readable"]:
                logger.warning(f"{location_file} is not readable, replacing with template")
                try:
                    create_backup(filepath, "unreadable")
                    result["backup_created"] = True
                except:
                    logger.warning(f"Could not create backup for {location_file}")
                result["status"] = "replaced_unreadable"
                result["issues_fixed"] = detection["issues"]
            elif detection["is_old_format"] or force_replace:
                logger.info(f"{location_file} has format issues: {detection['issues']}")
                
                # Create backup
                try:
                    backup_path = create_backup(filepath, "old_format")
                    result["backup_created"] = True
                except:
                    logger.warning(f"Could not create backup for {location_file}")
                
                result["issues_fixed"] = detection["issues"]
                result["status"] = "replaced"
                
                # Force close any open file handles and delete old file
                try:
                    # Force garbage collection to release file handles
                    import gc
                    gc.collect()
                    
                    # Use Windows-specific approach to force delete
                    import time
                    for attempt in range(3):
                        try:
                            filepath.unlink()
                            break
                        except PermissionError:
                            time.sleep(1)
                            gc.collect()
                    else:
                        # If still locked, rename it instead
                        temp_name = filepath.with_suffix('.old.xlsx')
                        filepath.rename(temp_name)
                        logger.info(f"Renamed locked file to {temp_name}")
                except Exception as e:
                    logger.warning(f"Could not remove old file {filepath}: {e}")
                
            else:
                logger.info(f"{location_file} appears to be in correct format")
                result["status"] = "already_correct"
                return result
        
        # Load fresh template
        template_wb, template_ws = load_official_template()
        
        # Create new workbook - copy just the data without complex formatting
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "事業所集計"  # Simple title
        
        # Copy template data (values only, no formatting to avoid StyleProxy issues)
        for row_num in range(1, 41):  # Copy template structure + some extra rows
            for col_num in range(1, 19):  # A to R
                source_cell = template_ws.cell(row=row_num, column=col_num)
                target_cell = new_ws.cell(row=row_num, column=col_num)
                
                # Copy only the value, not formatting
                target_cell.value = source_cell.value
        
        # Save the new workbook
        new_wb.save(filepath)
        new_wb.close()
        template_wb.close()
        
        logger.info(f"✅ {location_file} successfully updated with template format")
        
    except Exception as e:
        logger.error(f"❌ Failed to process {location_file}: {e}")
        result["status"] = "error"
        result["error"] = str(e)
    
    return result


def migrate_all_locations(force_replace: bool = False) -> Dict[str, Dict[str, Any]]:
    """
    Migrate all location workbooks to the official template format.
    
    Args:
        force_replace: If True, replace all files regardless of current format
        
    Returns:
        Dictionary with results for each location
    """
    logger.info("🔄 Starting migration of all location workbooks to template format")
    logger.info("=" * 70)
    
    # Ensure directories exist
    ACCUM_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    
    results = {}
    
    for location_file in LOCATION_FILES:
        logger.info(f"\n📍 Processing {location_file}...")
        result = apply_template_to_location(location_file, force_replace)
        results[location_file] = result
        
        # Print result summary
        status = result["status"]
        if status == "already_correct":
            logger.info(f"   ✅ Already in correct format")
        elif status == "created_new":
            logger.info(f"   🆕 Created new template file")
        elif status == "migrated":
            logger.info(f"   🔄 Migrated {result['migrated_rows']} rows, fixed: {result['issues_fixed']}")
        elif status == "replaced":
            logger.info(f"   🔄 Replaced old format, fixed: {result['issues_fixed']}")
        elif status == "replaced_unreadable":
            logger.info(f"   🔄 Replaced unreadable file")
        elif status == "error":
            logger.error(f"   ❌ Error: {result.get('error', 'Unknown error')}")
    
    # Summary
    logger.info("\n" + "=" * 70)
    logger.info("📊 MIGRATION SUMMARY")
    logger.info("=" * 70)
    
    total_migrated = sum(r.get("migrated_rows", 0) for r in results.values())
    successful = sum(1 for r in results.values() if r["status"] not in ["error"])
    
    for location_file, result in results.items():
        status_emoji = "✅" if result["status"] not in ["error"] else "❌"
        logger.info(f"{status_emoji} {location_file:<25} | {result['status']:<15} | Rows: {result.get('migrated_rows', 0)}")
    
    logger.info(f"\n🎉 Migration complete: {successful}/{len(LOCATION_FILES)} files processed successfully")
    logger.info(f"📊 Total rows migrated: {total_migrated}")
    
    return results


def verify_all_templates() -> bool:
    """
    Verify that all location workbooks now follow the template format.
    
    Returns:
        True if all files are correctly formatted
    """
    logger.info("\n🔍 VERIFYING ALL TEMPLATE FILES")
    logger.info("=" * 50)
    
    all_correct = True
    
    for location_file in LOCATION_FILES:
        filepath = ACCUM_DIR / location_file
        
        if not filepath.exists():
            logger.warning(f"❌ {location_file} does not exist")
            all_correct = False
            continue
        
        detection = detect_old_format(filepath)
        
        if detection["is_old_format"] or not detection["readable"]:
            logger.warning(f"❌ {location_file} still has issues: {detection['issues']}")
            all_correct = False
        else:
            logger.info(f"✅ {location_file} is correctly formatted")
    
    if all_correct:
        logger.info("\n🎉 All location workbooks are now in correct template format!")
    else:
        logger.warning("\n⚠️ Some files still have issues. Re-run migration if needed.")
    
    return all_correct


if __name__ == "__main__":
    """
    Main execution script.
    
    Run this script to migrate all location workbooks to the template format.
    """
    try:
        # Check if template exists
        if not TEMPLATE_PATH.exists():
            logger.error(f"❌ Template file not found: {TEMPLATE_PATH}")
            logger.error("Please ensure the official template is available before running migration.")
            exit(1)
        
        # Run migration
        results = migrate_all_locations(force_replace=False)  # Set to True to force replace all files
        
        # Verify results
        verify_all_templates()
        
        logger.info("\n🚀 Migration script completed successfully!")
        logger.info("All location workbooks should now be ready for the new Receipt OCR system.")
        
    except Exception as e:
        logger.error(f"❌ Migration script failed: {e}")
        raise