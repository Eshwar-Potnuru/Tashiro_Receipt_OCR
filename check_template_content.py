#!/usr/bin/env python3
"""Check the original template content to understand the issue."""

import openpyxl
from pathlib import Path

def check_template_content():
    """Check what's in the original template."""
    
    print("🔍 CHECKING ORIGINAL TEMPLATE CONTENT")
    print("=" * 50)
    
    template_path = Path('Template/事業所集計テーブル.xlsx')
    
    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        print(f"📊 Available sheets: {wb.sheetnames}")
        
        # Check the November 2025 sheet
        target_sheet = '2025年11月'
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
            print(f"\n📄 Sheet \"{target_sheet}\" content:")
            print(f"   Max row: {ws.max_row}")
            print(f"   Max col: {ws.max_column}")
            
            # Show first 10 rows
            print(f"\n📋 First 10 rows:")
            for row_num in range(1, min(11, ws.max_row + 1)):
                row_content = []
                for col in range(1, min(10, ws.max_column + 1)):
                    value = ws.cell(row=row_num, column=col).value
                    if value:
                        row_content.append(f"{chr(64+col)}:{str(value)[:15]}")
                
                if row_content:
                    print(f"   Row {row_num:2d}: {row_content}")
                else:
                    print(f"   Row {row_num:2d}: [EMPTY]")
            
            # Check if there's substantial content
            content_rows = 0
            for row_num in range(1, ws.max_row + 1):
                has_content = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row_num, column=col).value:
                        has_content = True
                        break
                if has_content:
                    content_rows += 1
            
            print(f"\n📊 Total rows with content: {content_rows}")
            
            if content_rows < 5:
                print("❌ PROBLEM: Original template appears to be mostly empty!")
            else:
                print("✅ Original template has substantial content")
        else:
            print(f"❌ Target sheet '{target_sheet}' not found")
            print(f"   Available: {wb.sheetnames}")
        
        wb.close()
    else:
        print("❌ Original template file not found!")
    
    # Also check a location file to see the difference
    print(f"\n🔍 CHECKING LOCATION FILE (Tokyo):")
    tokyo_file = Path('app/Data/accumulation/Tokyo_Accumulated.xlsx')
    if tokyo_file.exists():
        wb = openpyxl.load_workbook(tokyo_file)
        print(f"📊 Tokyo file sheets: {wb.sheetnames}")
        
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
            print(f"📄 Tokyo {target_sheet} max_row: {ws.max_row}")
            
            # Check first few rows
            content_rows = 0
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_content = []
                for col in range(1, 6):
                    value = ws.cell(row=row_num, column=col).value
                    if value:
                        row_content.append(f"{chr(64+col)}:{str(value)[:10]}")
                        content_rows += 1
                
                if row_content:
                    print(f"   Row {row_num}: {row_content}")
            
            if content_rows == 0:
                print("❌ CONFIRMED: Location file is completely empty!")
            
        wb.close()
    else:
        print("❌ Tokyo location file doesn't exist")

if __name__ == "__main__":
    check_template_content()