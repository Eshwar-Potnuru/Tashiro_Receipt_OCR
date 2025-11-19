#!/usr/bin/env python3
"""
COMPREHENSIVE SYSTEM CHECK - Verify all components are working correctly
"""

import openpyxl
import template_formatter
from pathlib import Path
from validators import get_available_locations

def comprehensive_system_check():
    """Perform a complete system verification."""
    
    print("🔍 COMPREHENSIVE SYSTEM CHECK")
    print("=" * 60)
    
    issues = []
    successes = []
    
    # 1. Check original template exists and has content
    print("1️⃣ CHECKING ORIGINAL TEMPLATE")
    print("-" * 30)
    
    template_path = Path('Template/事業所集計テーブル.xlsx')
    if not template_path.exists():
        issues.append("❌ Original template file missing")
        print("❌ Original template file missing")
    else:
        wb = openpyxl.load_workbook(template_path)
        target_sheet = '2025年11月'
        
        if target_sheet not in wb.sheetnames:
            issues.append(f"❌ Target sheet '{target_sheet}' missing in template")
            print(f"❌ Target sheet '{target_sheet}' missing")
        else:
            ws = wb[target_sheet]
            content_rows = sum(1 for row in range(1, ws.max_row + 1) 
                             if any(ws.cell(row, col).value for col in range(1, ws.max_column + 1)))
            
            if content_rows < 20:
                issues.append(f"❌ Template has insufficient content ({content_rows} rows)")
                print(f"❌ Template insufficient content: {content_rows} rows")
            else:
                successes.append(f"✅ Template has {content_rows} content rows")
                print(f"✅ Template verified: {content_rows} content rows")
        
        wb.close()
    
    # 2. Check all location files exist and have proper structure
    print("\n2️⃣ CHECKING LOCATION FILES")
    print("-" * 30)
    
    config = get_available_locations()
    locations = config.get('locations', [])
    accum_dir = Path('app/Data/accumulation')
    
    for location in locations:
        location_file = accum_dir / f"{location}_Accumulated.xlsx"
        
        if not location_file.exists():
            issues.append(f"❌ {location} file missing")
            print(f"❌ {location}: File missing")
            continue
        
        try:
            wb = openpyxl.load_workbook(location_file)
            
            if target_sheet not in wb.sheetnames:
                issues.append(f"❌ {location} missing target sheet")
                print(f"❌ {location}: Missing target sheet")
                wb.close()
                continue
            
            ws = wb[target_sheet]
            
            # Check key template elements
            row1_title = ws.cell(1, 1).value  # Should be "2025年"
            row4_header = ws.cell(4, 1).value  # Should be "支払日"
            row5_carryover = ws.cell(5, 3).value  # Should contain "繰越"
            
            template_intact = (
                row1_title and "2025年" in str(row1_title) and
                row4_header and "支払日" in str(row4_header) and
                row5_carryover and "繰越" in str(row5_carryover)
            )
            
            if template_intact:
                successes.append(f"✅ {location} template structure intact")
                print(f"✅ {location}: Template structure intact")
            else:
                issues.append(f"❌ {location} template structure damaged")
                print(f"❌ {location}: Template structure damaged")
                print(f"   Row 1: {row1_title}")
                print(f"   Row 4: {row4_header}")
                print(f"   Row 5: {row5_carryover}")
            
            wb.close()
            
        except Exception as e:
            issues.append(f"❌ {location} file error: {e}")
            print(f"❌ {location}: Error - {e}")
    
    # 3. Test the append functionality
    print("\n3️⃣ TESTING APPEND FUNCTIONALITY")
    print("-" * 30)
    
    test_data = {
        'date': '2024-11-19',
        'amount': 9876,
        'description': 'システムチェックテスト',
        'invoice_number': 'SYSTEM-CHECK-999'
    }
    
    test_operator = {'name': 'システムチェック'}
    
    # Test with Tokyo location
    try:
        result = template_formatter.append_to_formatted_template(
            test_data, 'Tokyo', test_operator
        )
        
        status = result.get('status')
        row_number = result.get('row_number')
        
        if status == 'success':
            if row_number and row_number >= 41:
                successes.append(f"✅ Append successful at row {row_number}")
                print(f"✅ Append test successful at row {row_number}")
                
                # Verify the data was actually written
                tokyo_file = accum_dir / "Tokyo_Accumulated.xlsx"
                wb = openpyxl.load_workbook(tokyo_file)
                ws = wb[target_sheet]
                
                written_invoice = ws.cell(row_number, 8).value  # Column H
                written_amount = ws.cell(row_number, 6).value   # Column F
                
                if (str(written_invoice) == 'SYSTEM-CHECK-999' and 
                    str(written_amount) == '9876'):
                    successes.append("✅ Data verification successful")
                    print("✅ Data written correctly")
                else:
                    issues.append("❌ Data not written correctly")
                    print(f"❌ Data mismatch - Invoice: {written_invoice}, Amount: {written_amount}")
                
                wb.close()
            else:
                issues.append(f"❌ Append at wrong row ({row_number})")
                print(f"❌ Append at wrong row: {row_number}")
        else:
            issues.append(f"❌ Append failed: {status}")
            print(f"❌ Append failed: {status}")
            
    except Exception as e:
        issues.append(f"❌ Append test error: {e}")
        print(f"❌ Append test error: {e}")
    
    # 4. Check row finding logic
    print("\n4️⃣ TESTING ROW FINDING LOGIC")
    print("-" * 30)
    
    try:
        tokyo_file = accum_dir / "Tokyo_Accumulated.xlsx"
        wb = openpyxl.load_workbook(tokyo_file)
        ws = wb[target_sheet]
        
        next_row = template_formatter.find_first_empty_row_at_bottom(ws)
        
        if next_row > 40:
            successes.append(f"✅ Row finding correct: {next_row}")
            print(f"✅ Row finding working: next empty row {next_row}")
        else:
            issues.append(f"❌ Row finding incorrect: {next_row}")
            print(f"❌ Row finding issue: found row {next_row} (expected >40)")
        
        wb.close()
        
    except Exception as e:
        issues.append(f"❌ Row finding error: {e}")
        print(f"❌ Row finding error: {e}")
    
    # 5. Check Japanese header mapping
    print("\n5️⃣ CHECKING JAPANESE HEADER MAPPING")
    print("-" * 30)
    
    expected_headers = [
        "支払日", "工番", "摘　　要", "担当者", "収入", "支出", "a", "インボイス",
        "勘定科目", "b", "10％税込額", "8％税込額", "非課税額", "税込合計", "c",
        "消費税10", "消費税8", "消費税計"
    ]
    
    try:
        test_row = template_formatter.prepare_japanese_row_values(
            test_data, 'Tokyo', test_operator
        )
        
        if len(test_row) == 18:
            successes.append("✅ Row mapping correct length")
            print("✅ Japanese mapping: 18 columns")
            
            # Check key mappings
            if test_row[0] == test_data['date']:  # Column A
                successes.append("✅ Date mapping correct")
                print("✅ Date mapping correct")
            else:
                issues.append("❌ Date mapping incorrect")
                print("❌ Date mapping incorrect")
                
            if test_row[5] == test_data['amount']:  # Column F
                successes.append("✅ Amount mapping correct")
                print("✅ Amount mapping correct")
            else:
                issues.append("❌ Amount mapping incorrect")
                print("❌ Amount mapping incorrect")
                
        else:
            issues.append(f"❌ Row mapping wrong length: {len(test_row)}")
            print(f"❌ Row mapping wrong length: {len(test_row)}")
            
    except Exception as e:
        issues.append(f"❌ Header mapping error: {e}")
        print(f"❌ Header mapping error: {e}")
    
    # 6. Check staff assignment
    print("\n6️⃣ CHECKING STAFF ASSIGNMENT")
    print("-" * 30)
    
    try:
        from accumulator import _get_staff_member_for_location
        
        staff_member = _get_staff_member_for_location('Tokyo', test_operator)
        
        if staff_member and len(staff_member) > 1:
            successes.append(f"✅ Staff assignment working: {staff_member}")
            print(f"✅ Staff assignment: {staff_member}")
        else:
            issues.append("❌ Staff assignment not working")
            print("❌ Staff assignment failed")
            
    except Exception as e:
        issues.append(f"❌ Staff assignment error: {e}")
        print(f"❌ Staff assignment error: {e}")
    
    # FINAL SUMMARY
    print("\n" + "=" * 60)
    print("📊 SYSTEM CHECK SUMMARY")
    print("=" * 60)
    
    print(f"\n✅ SUCCESSES ({len(successes)}):")
    for success in successes:
        print(f"   {success}")
    
    if issues:
        print(f"\n❌ ISSUES FOUND ({len(issues)}):")
        for issue in issues:
            print(f"   {issue}")
    else:
        print(f"\n🎉 NO ISSUES FOUND!")
    
    print(f"\n🎯 OVERALL STATUS:")
    if len(issues) == 0:
        print("✅ SYSTEM FULLY OPERATIONAL")
        print("✅ Template structure preserved")
        print("✅ Append functionality working")
        print("✅ All location files properly configured")
        print("✅ Ready for production use")
    elif len(issues) <= 2:
        print("⚠️  SYSTEM MOSTLY OPERATIONAL")
        print("🔧 Minor issues need attention")
    else:
        print("❌ SYSTEM NEEDS ATTENTION")
        print("🔧 Multiple issues require fixing")
    
    print("=" * 60)
    
    return len(issues) == 0

if __name__ == "__main__":
    all_good = comprehensive_system_check()
    if all_good:
        print("\n🚀 COMPREHENSIVE CHECK PASSED - SYSTEM READY!")
    else:
        print("\n🔧 ISSUES FOUND - REVIEW NEEDED")